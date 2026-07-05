# =============================================================================
# KAMIS Market Price Checker  —  v3.00 (faster rewrite)
# =============================================================================
#
# DESCRIPTION
# -----------
# Tkinter desktop application that scrapes KAMIS market prices for target
# products, filters data, and exports a formatted Excel workbook.
#
# This v3.00 build focuses on faster end-to-end scrape time while preserving
# the same GUI workflow and output structure (status UI, stop behavior,
# timestamped Excel output, and formatting).
#
# -----------------------------------------------------------------------------
# CHANGELOG
# -----------------------------------------------------------------------------
#
# v3.00  —  2026-04-28
#   - Duplicate from v2.10 for continued development and improvements.
#
# v2.10  —  2026-04-15
#   PERFORMANCE UPDATES
#   - Added parallel product scraping via ThreadPoolExecutor so "Dry Maize",
#     "Dry Onions", and "Chillies" can be fetched concurrently.
#   - Replaced pandas.read_html with a direct BeautifulSoup/lxml table parser
#     (parse_main_table_fast) to reduce HTML parsing overhead per page.
#   - Added one-time product mapping resolution per run
#     (resolve_product_mapping) instead of resolving product options per
#     product scrape.
#   - Increased HTTP connection pool capacity
#     (pool_connections=8, pool_maxsize=8) for better request reuse.
#   - Reduced inter-page delay default for faster pagination
#     (sleep_s=0.08 with jitter) while retaining polite request spacing.
#   - Kept retry/backoff and 429 Retry-After handling for reliability under
#     higher scrape throughput.
#
# v2.00  —  2026-04-15
#   - Baseline v2 duplicate from v1 with matching version labels and output
#     filename prefix updates.
# =============================================================================

import time
import sys
import subprocess
import random
import threading
import queue
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional
from collections import deque
from urllib.parse import urljoin, urlparse, parse_qsl, urlencode, urlunparse

import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter as tk
from tkinter import ttk, messagebox

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE = "https://kamis.kilimo.go.ke"
START_URL = "https://kamis.kilimo.go.ke/site/market"
TARGET_PRODUCTS = ["Dry Maize", "Dry Onions", "Chillies"]
TARGET_MARKET: Optional[str] = None
TARGET_PER_PAGE = 3000
OUTPUT_DIR = Path(r"c:\Users\kirwa\Documents\Farming and Agriculture\Kamis Data Results")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
}


@dataclass
class EtaEstimator:
    sample_size: int = 8
    samples: deque[float] = field(default_factory=lambda: deque(maxlen=8))
    _last_mark: float = field(default_factory=time.perf_counter)

    def mark_item_complete(self) -> None:
        now = time.perf_counter()
        delta = now - self._last_mark
        self._last_mark = now
        if delta > 0:
            self.samples.append(delta)

    def eta_seconds(self, completed: int, total: Optional[int]) -> Optional[float]:
        if total is None or completed <= 0 or total <= completed or not self.samples:
            return None
        avg_s = sum(self.samples) / len(self.samples)
        return avg_s * (total - completed)


def format_eta(seconds: Optional[float]) -> str:
    if seconds is None:
        return "Estimating..."
    seconds = max(0, int(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def format_elapsed(seconds: float) -> str:
    total_seconds = max(0, int(seconds))
    hours, rem = divmod(total_seconds, 3600)
    minutes, secs = divmod(rem, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def estimate_total_pages_from_soup(soup: BeautifulSoup) -> Optional[int]:
    max_page: Optional[int] = None
    for anchor in soup.find_all("a", href=True):
        parsed = urlparse(anchor["href"])
        params = dict(parse_qsl(parsed.query, keep_blank_values=True))
        raw = params.get("page")
        if raw and raw.isdigit():
            page_no = int(raw)
            if page_no > 0:
                max_page = page_no if max_page is None else max(max_page, page_no)
    return max_page


def is_excel_open() -> bool:
    if not sys.platform.startswith("win"):
        return False
    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq EXCEL.EXE"],
            capture_output=True,
            text=True,
            check=False,
        )
        return "excel.exe" in (result.stdout or "").lower()
    except Exception:
        return False


def close_excel_processes() -> bool:
    if not sys.platform.startswith("win"):
        return False
    try:
        result = subprocess.run(
            ["taskkill", "/IM", "EXCEL.EXE", "/F"],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def prompt_close_excel_if_open() -> None:
    if not is_excel_open():
        return

    print("[WARN] Microsoft Excel appears to be running.", flush=True)
    print("[WARN] If the target file is open in Excel, saving may fail.", flush=True)

    try:
        answer = input("Close Excel now? [y/N]: ").strip().lower()
    except EOFError:
        answer = "n"

    if answer in {"y", "yes"}:
        if close_excel_processes():
            print("[INFO] Excel process closed.", flush=True)
        else:
            print("[WARN] Could not close Excel automatically. Please close it manually.", flush=True)
    else:
        print("[INFO] Continuing without closing Excel.", flush=True)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    session.verify = False

    adapter = requests.adapters.HTTPAdapter(
        pool_connections=8,
        pool_maxsize=8,
        max_retries=0,
    )
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def fetch_with_retries(
    session: requests.Session,
    url: str,
    timeout: float,
    max_retries: int,
    retry_backoff_s: float,
    stop_event: Optional[threading.Event] = None,
) -> requests.Response:
    last_error = None

    for attempt in range(1, max_retries + 1):
        if stop_event and stop_event.is_set():
            raise RuntimeError("Scrape stopped by user.")

        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException as error:
            last_error = error
            force_wait_s = 0.0
            if isinstance(error, requests.exceptions.HTTPError) and error.response is not None:
                if error.response.status_code == 429:
                    retry_after = error.response.headers.get("Retry-After", "60")
                    try:
                        force_wait_s = float(retry_after)
                    except ValueError:
                        force_wait_s = 60.0

            if attempt < max_retries:
                wait_s = force_wait_s if force_wait_s else retry_backoff_s * (2 ** (attempt - 1))
                time.sleep(wait_s)

    raise RuntimeError(f"Request failed after {max_retries} attempts: {url}\nLast error: {last_error}") from last_error


def upsert_query_param(url: str, key: str, value: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[key] = value
    return urlunparse(parsed._replace(query=urlencode(query)))


def find_next_url(soup: BeautifulSoup, current_url: str) -> str | None:
    next_a = soup.find("a", string=lambda s: isinstance(s, str) and s.strip() == ">")
    if not next_a:
        return None

    href = next_a.get("href")
    if not href:
        return None

    return urljoin(current_url, href)


def parse_main_table_fast(html: str) -> pd.DataFrame:
    """
    Faster replacement for pd.read_html().
    Attempts to identify the main data table by preferring the table
    with the most rows and at least one header row.
    """
    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")
    if not tables:
        return pd.DataFrame()

    best_table = None
    best_score = -1

    for table in tables:
        rows = table.find_all("tr")
        header_cells = table.find_all("th")
        score = len(rows) * 10 + len(header_cells)
        if score > best_score:
            best_score = score
            best_table = table

    if best_table is None:
        return pd.DataFrame()

    rows = best_table.find_all("tr")
    if not rows:
        return pd.DataFrame()

    headers = []
    header_row = None

    for row in rows:
        ths = row.find_all("th")
        if ths:
            headers = [th.get_text(" ", strip=True) for th in ths]
            header_row = row
            break

    if not headers:
        first_cells = rows[0].find_all(["td", "th"])
        headers = [cell.get_text(" ", strip=True) for cell in first_cells]
        data_rows = rows[1:]
    else:
        data_rows = rows[rows.index(header_row) + 1 :]

    cleaned_headers = []
    seen = {}
    for h in headers:
        col = str(h).strip() or "Column"
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        cleaned_headers.append(col)

    records = []
    expected_cols = len(cleaned_headers)

    for row in data_rows:
        cells = row.find_all("td")
        if not cells:
            continue

        values = [cell.get_text(" ", strip=True) for cell in cells]

        if len(values) < expected_cols:
            values.extend([""] * (expected_cols - len(values)))
        elif len(values) > expected_cols:
            values = values[:expected_cols]

        records.append(values)

    if not records:
        return pd.DataFrame(columns=cleaned_headers)

    return pd.DataFrame(records, columns=cleaned_headers)


def is_effectively_empty_table(df: pd.DataFrame) -> bool:
    if df.empty:
        return True

    normalized = df.copy()
    for col in normalized.columns:
        if normalized[col].dtype == object:
            normalized[col] = normalized[col].astype(str).str.strip()

    normalized = normalized.replace(
        {"": pd.NA, "-": pd.NA, " - ": pd.NA, "nan": pd.NA, "None": pd.NA}
    )
    return normalized.dropna(how="all").empty


def filter_market_rows(df: pd.DataFrame, market_name: Optional[str]) -> pd.DataFrame:
    if not market_name:
        return df.copy()

    if "Market" not in df.columns:
        return df.iloc[0:0].copy()

    market_series = df["Market"].astype(str).str.strip().str.casefold()
    target = market_name.strip().casefold()
    return df.loc[market_series == target].copy()


def resolve_product_mapping(
    session: requests.Session,
    start_url: str,
    timeout: float,
    max_retries: int,
    retry_backoff_s: float,
    stop_event: Optional[threading.Event] = None,
) -> dict[str, str]:
    response = fetch_with_retries(
        session=session,
        url=start_url,
        timeout=timeout,
        max_retries=max_retries,
        retry_backoff_s=retry_backoff_s,
        stop_event=stop_event,
    )

    soup = BeautifulSoup(response.text, "lxml")
    select = soup.find("select", {"id": "selProduct"}) or soup.find("select", {"name": "product"})
    if not select:
        raise RuntimeError("Product filter select was not found on the market page.")

    mapping: dict[str, str] = {}
    for option in select.find_all("option"):
        label = option.get_text(" ", strip=True)
        value = (option.get("value") or "").strip()
        if label and value:
            mapping[label.strip().lower()] = value

    if not mapping:
        raise RuntimeError("No product options were found on the market page.")

    return mapping


def scrape_all_for_product(
    start_url: str,
    product_name: str,
    product_value: str,
    market_filter_name: Optional[str] = None,
    per_page: int = 3000,
    sleep_s: float = 0.08,
    max_pages: int = 2000,
    request_timeout: float = 40,
    max_retries: int = 3,
    retry_backoff_s: float = 2.0,
    stop_event: Optional[threading.Event] = None,
) -> pd.DataFrame:
    session = make_session()

    try:
        url = upsert_query_param(start_url, "product", product_value)
        url = upsert_query_param(url, "per_page", str(per_page))

        all_dfs: list[pd.DataFrame] = []
        seen: set[str] = set()
        pages_with_data = 0
        last_columns = None

        for page_idx in range(max_pages):
            if stop_event and stop_event.is_set():
                raise RuntimeError("Scrape stopped by user.")

            if url in seen:
                break
            seen.add(url)

            response = fetch_with_retries(
                session=session,
                url=url,
                timeout=request_timeout,
                max_retries=max_retries,
                retry_backoff_s=retry_backoff_s,
                stop_event=stop_event,
            )

            page_soup = BeautifulSoup(response.text, "lxml")
            df = parse_main_table_fast(response.text)

            if is_effectively_empty_table(df):
                break

            pages_with_data += 1
            last_columns = list(df.columns)

            market_df = filter_market_rows(df, market_filter_name)
            if not market_df.empty:
                market_df["Selected Product"] = product_name
                all_dfs.append(market_df)

            nxt = find_next_url(page_soup, url)
            if not nxt:
                break

            url = nxt

            if sleep_s > 0:
                jitter_s = sleep_s + random.uniform(0.0, sleep_s * 0.35)
                time.sleep(jitter_s)

        if not all_dfs:
            if pages_with_data == 0:
                raise RuntimeError(f"No table data was scraped for {product_name}.")
            return pd.DataFrame(columns=last_columns or [])

        out = pd.concat(all_dfs, ignore_index=True)
        dedup_cols = [
            c for c in out.columns
            if c in {"Date", "Market", "Commodity", "Wholesale Price", "Retail Price"}
        ]
        if dedup_cols:
            out = out.drop_duplicates(subset=dedup_cols)

        return out

    finally:
        session.close()


def add_wholesale_90kg_bag_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    new_col = "Wholesale 90Kg Bag (Maize)"
    if new_col in df.columns:
        return df

    wholesale_col = next((c for c in df.columns if "wholesale" in str(c).strip().lower()), None)
    if wholesale_col is None:
        return df

    wholesale_numeric = pd.to_numeric(
        df[wholesale_col]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.extract(r"([\d.]+)", expand=False),
        errors="coerce",
    )

    wholesale_90kg = wholesale_numeric * 90

    out = df.copy()
    insert_at = out.columns.get_loc(wholesale_col) + 1
    out.insert(insert_at, new_col, wholesale_90kg)
    return out


def save_to_excel_with_table(df: pd.DataFrame, output_path: str) -> None:
    df.to_excel(output_path, index=False, sheet_name="Data")

    wb = load_workbook(output_path)
    ws = wb["Data"]

    if ws.max_column == 0:
        wb.save(output_path)
        return

    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row

    if last_row >= 2:
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="KamisData", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    kes_columns = {"Wholesale 90Kg Bag (Maize)", "Wholesale", "Retail", "Wholesale Price", "Retail Price"}
    kes_format = '"KES" #,##0.00'

    def parse_kes_number(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if not text:
            return None

        cleaned = (
            text.replace("KES", "")
            .replace("Ksh", "")
            .replace("ksh", "")
            .replace(",", "")
            .strip()
        )
        try:
            return float(cleaned)
        except ValueError:
            return None

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    for header in kes_columns:
        col_idx = header_map.get(header)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            numeric_value = parse_kes_number(cell.value)
            if numeric_value is None:
                continue
            cell.value = numeric_value
            cell.number_format = kes_format

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


class MaizePriceCheckerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("KAMIS Market Price Checker v3.00")
        self.root.geometry("920x500")
        self.root.minsize(860, 460)

        self.msg_queue: queue.Queue[dict] = queue.Queue()
        self.worker_thread: Optional[threading.Thread] = None
        self.stop_event = threading.Event()
        self.pending_close = False
        self.run_started_at = 0.0
        self.last_output_uri: Optional[str] = None
        self.spinner_frames = ["|", "/", "-", "\\"]
        self.spinner_index = 0
        self.spinner_job: Optional[str] = None
        self.elapsed_job: Optional[str] = None

        self.state_var = tk.StringVar(value="Idle")
        self.status_var = tk.StringVar(value="Ready")
        self.items_var = tk.StringVar(value="0 / ?")
        self.elapsed_var = tk.StringVar(value="00:00")
        self.spinner_var = tk.StringVar(value="-")

        self._build_ui()
        self._set_idle_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.after(150, self.process_queue)

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=14)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="KAMIS Market Price Checker v3.00", font=("Segoe UI", 15, "bold")).pack(anchor="w")

        button_row = ttk.Frame(main)
        button_row.pack(fill="x", pady=(12, 10))

        self.start_btn = ttk.Button(button_row, text="Start", command=self.start_scrape)
        self.start_btn.pack(side="left")

        self.stop_btn = ttk.Button(button_row, text="Stop", command=self.stop_scrape)
        self.stop_btn.pack(side="left", padx=(8, 0))

        ttk.Frame(button_row).pack(side="left", fill="x", expand=True)
        self.close_btn = ttk.Button(button_row, text="Close", command=self.close_app)
        self.close_btn.pack(side="right")

        status_frame = ttk.LabelFrame(main, text="Status", padding=10)
        status_frame.pack(fill="x")

        ttk.Label(status_frame, textvariable=self.state_var).pack(anchor="w")
        ttk.Label(status_frame, textvariable=self.status_var).pack(anchor="w", pady=(4, 0))

        spinner_row = ttk.Frame(main)
        spinner_row.pack(fill="x", pady=(14, 8))
        ttk.Label(spinner_row, text="Activity:").pack(side="left")
        ttk.Label(spinner_row, textvariable=self.spinner_var, font=("Consolas", 16, "bold")).pack(side="left", padx=(8, 0))

        metrics_row = ttk.Frame(main)
        metrics_row.pack(fill="x")

        ttk.Label(metrics_row, text="Items:").grid(row=0, column=0, sticky="w")
        ttk.Label(metrics_row, textvariable=self.items_var).grid(row=0, column=1, sticky="w", padx=(6, 0))
        ttk.Label(metrics_row, text="Elapsed:").grid(row=0, column=2, sticky="w", padx=(24, 0))
        ttk.Label(metrics_row, textvariable=self.elapsed_var).grid(row=0, column=3, sticky="w", padx=(6, 0))

        result_frame = ttk.LabelFrame(main, text="Result", padding=10)
        result_frame.pack(fill="both", expand=True, pady=(12, 0))

        self.result_text = tk.Text(result_frame, height=8, wrap="none")
        self.result_text.pack(side="left", fill="both", expand=True)
        yscroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_text.yview)
        yscroll.pack(side="right", fill="y")
        self.result_text.configure(yscrollcommand=yscroll.set)
        self.result_text.configure(state="disabled")

        link_row = ttk.Frame(main)
        link_row.pack(fill="x", pady=(10, 0))

        ttk.Label(link_row, text="Produced file:").pack(side="left")
        self.open_link = tk.Label(
            link_row,
            text="No file yet",
            fg="gray",
            cursor="arrow",
            anchor="w",
        )
        self.open_link.pack(side="left", padx=(8, 0), fill="x", expand=True)
        self.open_link.bind("<Button-1>", self.on_open_link_clicked)

    def set_result_text(self, text: str) -> None:
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", tk.END)
        self.result_text.insert(tk.END, text)
        self.result_text.configure(state="disabled")

    def _set_idle_ui(self) -> None:
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.stop_elapsed_timer()
        self.stop_spinner()

    def tick_elapsed(self) -> None:
        if self.run_started_at > 0:
            elapsed = time.perf_counter() - self.run_started_at
            self.elapsed_var.set(format_elapsed(elapsed))
        self.elapsed_job = self.root.after(200, self.tick_elapsed)

    def start_elapsed_timer(self) -> None:
        self.stop_elapsed_timer()
        self.elapsed_var.set("00:00")
        self.elapsed_job = self.root.after(200, self.tick_elapsed)

    def stop_elapsed_timer(self) -> None:
        if self.elapsed_job is not None:
            self.root.after_cancel(self.elapsed_job)
            self.elapsed_job = None

    def tick_spinner(self) -> None:
        self.spinner_var.set(self.spinner_frames[self.spinner_index])
        self.spinner_index = (self.spinner_index + 1) % len(self.spinner_frames)
        self.spinner_job = self.root.after(120, self.tick_spinner)

    def start_spinner(self) -> None:
        if self.spinner_job is None:
            self.spinner_index = 0
            self.tick_spinner()

    def stop_spinner(self) -> None:
        if self.spinner_job is not None:
            self.root.after_cancel(self.spinner_job)
            self.spinner_job = None
        self.spinner_var.set("-")

    def set_open_file_link(self, file_uri: Optional[str], text: str) -> None:
        self.last_output_uri = file_uri
        if file_uri:
            self.open_link.config(text=text, fg="#005A9C", cursor="hand2")
        else:
            self.open_link.config(text=text, fg="gray", cursor="arrow")

    def on_open_link_clicked(self, _event=None) -> None:
        if not self.last_output_uri:
            return
        try:
            webbrowser.open(self.last_output_uri)
        except Exception as error:
            messagebox.showerror("Open file", f"Could not open file link.\n\n{error}")

    def close_app(self) -> None:
        self.on_close()

    def start_scrape(self) -> None:
        if self.worker_thread and self.worker_thread.is_alive():
            return

        self.stop_event.clear()
        self.pending_close = False
        self.run_started_at = time.perf_counter()
        self.set_result_text("")
        self.set_open_file_link(None, "No file yet")
        self.state_var.set("Running")
        self.status_var.set("Starting scraper...")
        self.items_var.set("0 / ?")
        self.elapsed_var.set("00:00")

        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.start_elapsed_timer()
        self.start_spinner()

        self.worker_thread = threading.Thread(target=self.worker_run_scraper, daemon=True)
        self.worker_thread.start()

    def stop_scrape(self) -> None:
        if not self.worker_thread or not self.worker_thread.is_alive():
            return
        self.state_var.set("Stopping")
        self.status_var.set("Waiting for safe stop...")
        self.stop_event.set()
        self.stop_btn.config(state="disabled")

    def worker_run_scraper(self) -> None:
        def emit(**payload) -> None:
            self.msg_queue.put(payload)

        try:
            bootstrap_session = make_session()
            try:
                emit(status="Loading product mapping...", completed=0, total=len(TARGET_PRODUCTS))
                product_mapping = resolve_product_mapping(
                    session=bootstrap_session,
                    start_url=START_URL,
                    timeout=40,
                    max_retries=3,
                    retry_backoff_s=2.0,
                    stop_event=self.stop_event,
                )
            finally:
                bootstrap_session.close()

            normalized_map = {k.strip().lower(): v for k, v in product_mapping.items()}

            missing = [p for p in TARGET_PRODUCTS if p.strip().lower() not in normalized_map]
            if missing:
                raise RuntimeError(f"Missing product filters on KAMIS page: {', '.join(missing)}")

            all_product_frames: list[pd.DataFrame] = []
            total_products = len(TARGET_PRODUCTS)
            completed_products = 0

            emit(status="Scraping products in parallel...", completed=0, total=total_products)

            with ThreadPoolExecutor(max_workers=min(3, total_products)) as executor:
                future_to_product = {
                    executor.submit(
                        scrape_all_for_product,
                        start_url=START_URL,
                        product_name=product_name,
                        product_value=normalized_map[product_name.strip().lower()],
                        market_filter_name=TARGET_MARKET,
                        per_page=TARGET_PER_PAGE,
                        sleep_s=0.08,
                        max_pages=2000,
                        request_timeout=40,
                        max_retries=3,
                        retry_backoff_s=2.0,
                        stop_event=self.stop_event,
                    ): product_name
                    for product_name in TARGET_PRODUCTS
                }

                for future in as_completed(future_to_product):
                    if self.stop_event.is_set():
                        self.msg_queue.put({"kind": "stopped"})
                        return

                    product_name = future_to_product[future]
                    product_df = future.result()

                    if not product_df.empty:
                        all_product_frames.append(product_df)

                    completed_products += 1
                    emit(
                        status=f"Completed {product_name}",
                        completed=completed_products,
                        total=total_products,
                    )

            if all_product_frames:
                df = pd.concat(all_product_frames, ignore_index=True)
            else:
                df = pd.DataFrame()

            if self.stop_event.is_set():
                self.msg_queue.put({"kind": "stopped"})
                return

            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
                current_year = datetime.now().year
                valid_years = {current_year - year_offset for year_offset in range(4)}
                valid_mask = df["Date"].notna() & df["Date"].dt.year.isin(valid_years)
                df = df.loc[valid_mask].copy()

            df = df.drop(columns=["Grade", "Sex", "Selected Product"], errors="ignore")
            df = add_wholesale_90kg_bag_column(df)

            now = datetime.now()
            timestamp = f"{now.month}-{now.day}-{now.year}  {now.strftime('%I.%M.%S %p')}"
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_file = OUTPUT_DIR / f"kamis_market_price_checker_v3.00_{timestamp}.xlsx"

            emit(status="Writing Excel file...", completed=total_products, total=total_products)
            save_to_excel_with_table(df, str(output_file))

            self.msg_queue.put(
                {
                    "kind": "completed",
                    "rows": len(df),
                    "file_link": output_file.resolve().as_uri(),
                    "file_path": str(output_file),
                    "elapsed": time.perf_counter() - self.run_started_at,
                }
            )

        except Exception as error:
            if self.stop_event.is_set():
                self.msg_queue.put({"kind": "stopped"})
            else:
                self.msg_queue.put({"kind": "error", "message": str(error)})

    def process_queue(self) -> None:
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                kind = msg.get("kind")

                if kind == "completed":
                    self.state_var.set("Completed")
                    self.status_var.set("Scrape completed successfully")
                    self.set_open_file_link(msg["file_link"], msg["file_path"])
                    self.set_result_text(
                        f"Rows: {msg['rows']:,}\n"
                        f"File: {msg['file_path']}\n"
                        f"Open file: {msg['file_link']}\n"
                        f"Elapsed: {format_elapsed(msg['elapsed'])}"
                    )
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                elif kind == "stopped":
                    self.state_var.set("Stopped")
                    self.status_var.set("Scrape stopped safely")
                    self.set_open_file_link(None, "No file produced")
                    self.set_result_text("No output file was produced because the run was stopped.")
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                elif kind == "error":
                    self.state_var.set("Idle")
                    self.status_var.set("Failed")
                    self.set_open_file_link(None, "No file produced")
                    self.set_result_text(f"Error: {msg.get('message', 'Unknown error')}")
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                else:
                    completed = int(msg.get("completed", 0))
                    total = msg.get("total")
                    self.status_var.set(msg.get("status", "Running..."))

                    if total and total > 0:
                        self.items_var.set(f"{completed} / {total}")
                    else:
                        self.items_var.set(f"{completed} / ?")

        except queue.Empty:
            pass
        finally:
            self.root.after(150, self.process_queue)

    def on_close(self) -> None:
        running = self.worker_thread and self.worker_thread.is_alive()
        if running:
            should_stop = messagebox.askyesno(
                "Stop scraper?",
                "A scrape is in progress. Stop safely and close the app?",
            )
            if not should_stop:
                return
            self.pending_close = True
            self.stop_scrape()
            self.state_var.set("Stopping")
            self.status_var.set("Stopping before close...")
            return
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = MaizePriceCheckerApp(root)
    root.mainloop()
