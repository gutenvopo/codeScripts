import time
import sys
import subprocess
from io import StringIO
from pathlib import Path
import pandas as pd
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, parse_qsl, urlencode, urlunparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = "https://kamis.kilimo.go.ke"
START_URL = "https://kamis.kilimo.go.ke/site/market"  # or any filtered URL you have
APP_VERSION = "2.00"
TARGET_PRODUCT = "Dry Maize"
TARGET_MARKET = "Eldoret Main"
TARGET_PER_PAGE = 3000


def is_excel_open() -> bool:
    """Detect whether Microsoft Excel is currently running (Windows only)."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq EXCEL.EXE"],
            capture_output=True,
            text=True,
            check=False,
        )
        return "excel.exe" in (result.stdout or "").lower()
    except Exception:
        return False


def close_excel_processes() -> bool:
    """Attempt to close all Excel processes. Returns True when successful."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["taskkill", "/IM", "EXCEL.EXE", "/F"],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def prompt_close_excel_if_open() -> None:
    """Prompt user to close Excel before writing CSV to avoid file-lock issues."""
    if not is_excel_open():
        return

    print("[WARN] Microsoft Excel appears to be running.", flush=True)
    print("[WARN] If the CSV file is open in Excel, saving may fail.", flush=True)

    try:
        answer = input("Close Excel now? [y/N]: ").strip().lower()
    except EOFError:
        answer = "n"

    if answer in {"y", "yes"}:
        if close_excel_processes():
            print("[INFO] Excel process closed.", flush=True)
        else:
            print("[WARN] Could not close Excel automatically. Please close it manually.", flush=True)
    else:
        print("[INFO] Continuing without closing Excel.", flush=True)


def convert_csv_to_styled_excel(csv_path: Path, excel_path: Path) -> None:
    """Convert CSV to Excel and apply Green Table Style Medium 7."""
    data = pd.read_csv(csv_path)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name="Data")

    wb = load_workbook(excel_path)
    ws = wb["Data"]
    max_row = ws.max_row
    max_col = ws.max_column

    # Excel table requires at least header + one data row.
    if max_col > 0 and max_row >= 2:
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="KamisDataTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(excel_path)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
}

def extract_table(html: str) -> pd.DataFrame:
    # KAMIS renders a single main table in this page; read_html usually finds it.
    tables = pd.read_html(StringIO(html))
    df = tables[-1].copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def is_effectively_empty_table(df: pd.DataFrame) -> bool:
    """Return True when a parsed table has no meaningful data rows."""
    if df.empty:
        return True

    normalized = df.copy()
    for col in normalized.columns:
        if normalized[col].dtype == object:
            normalized[col] = normalized[col].astype(str).str.strip()

    normalized = normalized.replace({"": pd.NA, "-": pd.NA, " - ": pd.NA, "nan": pd.NA, "None": pd.NA})
    return normalized.dropna(how="all").empty

def filter_market_rows(df: pd.DataFrame, market_name: str) -> pd.DataFrame:
    if "Market" not in df.columns:
        return df.iloc[0:0].copy()
    market_series = df["Market"].astype(str).str.strip().str.casefold()
    target = market_name.strip().casefold()
    return df.loc[market_series == target].copy()

def find_next_url(html: str, current_url: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")

    # The ">" next button is an <a> whose visible text is ">"
    next_a = soup.find("a", string=lambda s: isinstance(s, str) and s.strip() == ">")
    if not next_a:
        return None

    href = next_a.get("href")
    if not href:
        return None

    # Make it absolute
    return urljoin(current_url, href)

def upsert_query_param(url: str, key: str, value: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[key] = value
    return urlunparse(parsed._replace(query=urlencode(query)))

def resolve_product_value(
    session: requests.Session,
    start_url: str,
    product_name: str,
    timeout: float,
    max_retries: int,
    retry_backoff_s: float,
) -> str:
    """Resolve product option value from the market page (e.g., Dry Maize -> 1)."""
    response = None
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            response = session.get(start_url, timeout=timeout)
            response.raise_for_status()
            break
        except requests.exceptions.RequestException as error:
            last_error = error
            if attempt < max_retries:
                time.sleep(retry_backoff_s * (2 ** (attempt - 1)))

    if response is None:
        raise RuntimeError("Unable to load market filter options.") from last_error

    soup = BeautifulSoup(response.text, "html.parser")
    select = soup.find("select", {"id": "selProduct"}) or soup.find("select", {"name": "product"})
    if not select:
        raise RuntimeError("Product filter select was not found on the market page.")

    target = product_name.strip().lower()
    for option in select.find_all("option"):
        label = option.get_text(" ", strip=True).lower()
        value = (option.get("value") or "").strip()
        if label == target and value:
            return value

    raise RuntimeError(f"Product filter option not found: {product_name}")

def scrape_all(
    start_url: str,
    sleep_s: float = 0.4,
    max_pages: int = 2000,
    verbose: bool = True,
    request_timeout: float = 40,
    max_retries: int = 3,
    retry_backoff_s: float = 2.0,
    product_filter_name: str = "Dry Maize",
    market_filter_name: str = "Eldoret Main",
    per_page: int = 3000,
) -> pd.DataFrame:
    session = requests.Session()
    session.headers.update(HEADERS)

    filter_value = resolve_product_value(
        session=session,
        start_url=start_url,
        product_name=product_filter_name,
        timeout=request_timeout,
        max_retries=max_retries,
        retry_backoff_s=retry_backoff_s,
    )

    url = upsert_query_param(start_url, "product", filter_value)
    url = upsert_query_param(url, "per_page", str(per_page))
    all_dfs = []
    seen = set()
    started_at = time.perf_counter()
    pages_with_data = 0
    last_columns = None

    if verbose:
        print(
            "[START] Scrape started "
            f"| max_pages={max_pages} | sleep_s={sleep_s} "
            f"| timeout={request_timeout}s | max_retries={max_retries} "
            f"| product={product_filter_name}({filter_value}) "
            f"| market={market_filter_name} | per_page={per_page}",
            flush=True,
        )

    for page_idx in range(max_pages):
        page_no = page_idx + 1
        url = upsert_query_param(url, "product", filter_value)
        url = upsert_query_param(url, "per_page", str(per_page))
        if url in seen:
            if verbose:
                print(f"[STOP] Revisited URL detected at page {page_no}; stopping.", flush=True)
            break
        seen.add(url)

        if verbose:
            print(f"[PAGE {page_no}] Fetching URL: {url}", flush=True)

        req_start = time.perf_counter()
        r = None
        last_error = None
        for attempt in range(1, max_retries + 1):
            try:
                if verbose and attempt > 1:
                    print(f"[PAGE {page_no}] Retry {attempt}/{max_retries}", flush=True)
                r = session.get(url, timeout=request_timeout)
                r.raise_for_status()
                break
            except requests.exceptions.RequestException as error:
                last_error = error
                if attempt < max_retries:
                    wait_s = retry_backoff_s * (2 ** (attempt - 1))
                    if verbose:
                        print(
                            f"[PAGE {page_no}] Request failed ({type(error).__name__}): {error}",
                            flush=True,
                        )
                        print(f"[PAGE {page_no}] Backing off for {wait_s:.2f}s before retry", flush=True)
                    time.sleep(wait_s)
                else:
                    if verbose:
                        print(
                            f"[PAGE {page_no}] Request failed after {max_retries} attempts; stopping scrape.",
                            flush=True,
                        )

        if r is None:
            if all_dfs:
                if verbose:
                    print(f"[STOP] Returning partial data after page {page_no - 1} due to request failures.", flush=True)
                break
            raise RuntimeError("Unable to fetch the first page.") from last_error

        req_seconds = time.perf_counter() - req_start

        if verbose:
            print(
                f"[PAGE {page_no}] HTTP {r.status_code} in {req_seconds:.2f}s | bytes={len(r.text):,}",
                flush=True,
            )

        parse_start = time.perf_counter()
        df = extract_table(r.text)
        parse_seconds = time.perf_counter() - parse_start

        if is_effectively_empty_table(df):
            if verbose:
                print(f"[STOP] Empty table detected on page {page_no}; stopping.", flush=True)
            break

        pages_with_data += 1
        last_columns = list(df.columns)

        market_df = filter_market_rows(df, market_filter_name)
        if verbose:
            print(
                f"[PAGE {page_no}] Market filter kept {len(market_df):,}/{len(df):,} rows for '{market_filter_name}'",
                flush=True,
            )

        if not market_df.empty:
            all_dfs.append(market_df)

        if verbose:
            print(
                f"[PAGE {page_no}] Parsed table in {parse_seconds:.2f}s | rows={len(df):,}",
                flush=True,
            )

        next_start = time.perf_counter()
        nxt = find_next_url(r.text, url)
        next_seconds = time.perf_counter() - next_start

        if verbose:
            print(f"[PAGE {page_no}] Next-link parse in {next_seconds:.2f}s", flush=True)

        if not nxt:
            if verbose:
                print(f"[STOP] No next page after page {page_no}; stopping.", flush=True)
            break

        url = nxt
        if verbose:
            print(f"[PAGE {page_no}] Sleeping for {sleep_s:.2f}s", flush=True)
        time.sleep(sleep_s)

    if not all_dfs:
        if pages_with_data == 0:
            raise RuntimeError("No table data was scraped.")
        out = pd.DataFrame(columns=last_columns or [])
        if verbose:
            total_seconds = time.perf_counter() - started_at
            print(
                f"[DONE] Pages={pages_with_data:,} | Rows=0 | total={total_seconds:.2f}s",
                flush=True,
            )
        return out

    concat_start = time.perf_counter()
    out = pd.concat(all_dfs, ignore_index=True)
    concat_seconds = time.perf_counter() - concat_start

    if verbose:
        total_seconds = time.perf_counter() - started_at
        print(
            f"[DONE] Pages={pages_with_data:,} | Rows={len(out):,} | concat={concat_seconds:.2f}s | total={total_seconds:.2f}s",
            flush=True,
        )

    return out

if __name__ == "__main__":
    app_start = time.perf_counter()
    try:
        print(f"[INFO] Running Maize_Price_Checker v{APP_VERSION}", flush=True)
        prompt_close_excel_if_open()

        # Use larger page size to reduce pagination requests while preserving result fidelity.
        df = scrape_all(
            START_URL,
            product_filter_name=TARGET_PRODUCT,
            market_filter_name=TARGET_MARKET,
            per_page=TARGET_PER_PAGE,
        )

        # Optional cleanups
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

        csv_path = Path("kamis_market_dump.csv")
        excel_path = csv_path.with_suffix(".xlsx")

        df.to_csv(csv_path, index=False)
        convert_csv_to_styled_excel(csv_path, excel_path)

        print(df.head(3))
        print(f"\nRows: {len(df):,}")
        print(f"Saved CSV: {csv_path}")
        print(f"Saved Excel: {excel_path} (table style: Green Table Style Medium 7)")
    except KeyboardInterrupt:
        print("[CANCELLED] Scrape interrupted by user.", flush=True)
        sys.exit(130)
    except Exception as error:
        print(f"[ERROR] Scrape failed: {error}", flush=True)
        sys.exit(1)
    finally:
        print(f"Elapsed: {time.perf_counter() - app_start:.2f}s")