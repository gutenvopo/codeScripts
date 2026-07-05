# =============================================================================
# Onion Price Checker  —  v1.00
# =============================================================================
#
# DESCRIPTION
# -----------
# Scrapes onion wholesale and retail price data from the Kenya Agricultural
# Market Information System (KAMIS) website at https://kamis.kilimo.go.ke.
#
# The script targets a specific product ("Dry Onions") and market ("Eldoret Main"),
# filters results to the current year and the previous year only, removes
# unnecessary columns (Grade, Sex), and saves the cleaned data to a
# timestamped Excel spreadsheet formatted with an Orange Table Style Medium 7
# table, auto-fitted column widths, and saved to a user-defined output folder.
#
# -----------------------------------------------------------------------------
# CHANGELOG
# -----------------------------------------------------------------------------
#
# v1.00  —  2026-04-12
#   - Onion edition based on Maize_Price_Checker_v5_00 with full GUI support.
#
# v4.00  —  2026-04-12
#   OPTIMIZATIONS (performance, reliability, polite scraping)
#   - Added random jitter (up to 50 % of sleep_s) between page requests so
#     traffic does not look like a clock-driven bot.
#   - Added explicit 429 rate-limit handling: honours the server's Retry-After
#     header (defaults to 60 s if absent) before retrying, instead of the
#     normal short backoff.
#   - HTML is now parsed only once per page using lxml (BS4); the resulting
#     soup object is reused for next-link detection, eliminating a second full
#     parse that was happening silently on every page.
#   - Switched BS4 HTML parser from the built-in html.parser to lxml
#     (3–5× faster) in both scrape_all and resolve_product_value.
#   - Removed redundant upsert_query_param calls from inside the pagination
#     loop (URL parameters that never change were being rebuilt every page).
#   - Added HTTPAdapter with explicit connection pool settings so the TCP
#     connection to the server stays alive across all page requests instead
#     of re-handshaking each time.
#   - Added post-scrape deduplication keyed on Date, Market, Commodity,
#     Wholesale Price, and Retail Price to silently remove any rows duplicated
#     by pagination overlap; logged when rows are removed.
#   - Expanded request headers to include Accept, Accept-Language, and
#     Connection: keep-alive so requests look more like a real browser.
#
# v3.00  —  2026-04-12
#   - Changed output format from CSV to Excel (.xlsx).
#   - Applied Orange "Table Style Medium 7" table formatting via openpyxl.
#   - Added Excel auto-fit column widths (capped at 60 characters).
#   - Set output directory to:
#       C:\Users\kirwa\Documents\Farming and Agriculture\Kamis Data Results
#   - Output directory is created automatically if it does not exist.
#   - Timestamped filename format changed to:  M-D-YYYY  HH.MM.SS AM/PM
#   - Removed Grade and Sex columns from the final spreadsheet.
#   - After saving, prints a clickable file:// URI link to the created file.
#
# v2.00  —  2026-04-12
#   - Added year filtering: only data from the current year and the previous
#     year is kept; older rows are discarded before saving.
#   - Added SSL certificate verification bypass (self-signed cert on server)
#     with urllib3 InsecureRequestWarning suppressed.
#   - Improved error messages to include the specific error type and attempt
#     count when the initial option-resolve request fails.
#
# v1.00  —  initial release
#   - Scrapes all pages of the KAMIS market table for Dry Maize / Eldoret Main.
#   - Resolves the product filter value dynamically from the page HTML.
#   - Paginates until no next-page link is found or max_pages is reached.
#   - Filters rows to the target market before collecting.
#   - Handles transient network failures with configurable exponential backoff.
#   - Detects and stops on empty or duplicate pages.
#   - Saves output to kamis_market_dump.csv.
#   - Detects open Excel processes before saving and offers to close them.
#
# =============================================================================

import time
import sys
import subprocess
import random
import threading
import queue
import webbrowser
from io import StringIO
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter as tk
from tkinter import ttk, messagebox
from urllib.parse import urljoin, urlparse, parse_qsl, urlencode, urlunparse

# Suppress SSL warnings for self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE = "https://kamis.kilimo.go.ke"
START_URL = "https://kamis.kilimo.go.ke/site/market"  # or any filtered URL you have
TARGET_PRODUCT = "Dry Onions"
TARGET_MARKET = "Eldoret Main"
TARGET_PER_PAGE = 3000
OUTPUT_DIR = Path(r"c:\Users\kirwa\Documents\Farming and Agriculture\Kamis Data Results")


@dataclass
class EtaEstimator:
    """Rolling-average ETA estimator to avoid jumpy countdowns."""
    sample_size: int = 8
    samples: deque[float] = field(default_factory=lambda: deque(maxlen=8))
    _last_mark: float = field(default_factory=time.perf_counter)

    def mark_item_complete(self) -> None:
        now = time.perf_counter()
        delta = now - self._last_mark
        self._last_mark = now
        if delta > 0:
            self.samples.append(delta)

    def eta_seconds(self, completed: int, total: Optional[int]) -> Optional[float]:
        if total is None or completed <= 0 or total <= completed or not self.samples:
            return None
        avg_s = sum(self.samples) / len(self.samples)
        return avg_s * (total - completed)


def format_eta(seconds: Optional[float]) -> str:
    if seconds is None:
        return "Estimating..."
    seconds = max(0, int(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def estimate_total_pages_from_soup(soup: BeautifulSoup) -> Optional[int]:
    """Best-effort total-page detection from pagination links."""
    max_page: Optional[int] = None
    for anchor in soup.find_all("a", href=True):
        parsed = urlparse(anchor["href"])
        params = dict(parse_qsl(parsed.query, keep_blank_values=True))
        raw = params.get("page")
        if raw and raw.isdigit():
            page_no = int(raw)
            if page_no > 0:
                max_page = page_no if max_page is None else max(max_page, page_no)
    return max_page


def is_excel_open() -> bool:
    """Detect whether Microsoft Excel is currently running (Windows only)."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq EXCEL.EXE"],
            capture_output=True,
            text=True,
            check=False,
        )
        return "excel.exe" in (result.stdout or "").lower()
    except Exception:
        return False


def close_excel_processes() -> bool:
    """Attempt to close all Excel processes. Returns True when successful."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["taskkill", "/IM", "EXCEL.EXE", "/F"],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def prompt_close_excel_if_open() -> None:
    """Prompt user to close Excel before writing CSV to avoid file-lock issues."""
    if not is_excel_open():
        return

    print("[WARN] Microsoft Excel appears to be running.", flush=True)
    print("[WARN] If the CSV file is open in Excel, saving may fail.", flush=True)

    try:
        answer = input("Close Excel now? [y/N]: ").strip().lower()
    except EOFError:
        answer = "n"

    if answer in {"y", "yes"}:
        if close_excel_processes():
            print("[INFO] Excel process closed.", flush=True)
        else:
            print("[WARN] Could not close Excel automatically. Please close it manually.", flush=True)
    else:
        print("[INFO] Continuing without closing Excel.", flush=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
}

def extract_table(html: str) -> pd.DataFrame:
    # KAMIS renders a single main table in this page; read_html usually finds it.
    tables = pd.read_html(StringIO(html))
    df = tables[-1].copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def is_effectively_empty_table(df: pd.DataFrame) -> bool:
    """Return True when a parsed table has no meaningful data rows."""
    if df.empty:
        return True

    normalized = df.copy()
    for col in normalized.columns:
        if normalized[col].dtype == object:
            normalized[col] = normalized[col].astype(str).str.strip()

    normalized = normalized.replace({"": pd.NA, "-": pd.NA, " - ": pd.NA, "nan": pd.NA, "None": pd.NA})
    return normalized.dropna(how="all").empty

def filter_market_rows(df: pd.DataFrame, market_name: str) -> pd.DataFrame:
    if "Market" not in df.columns:
        return df.iloc[0:0].copy()
    market_series = df["Market"].astype(str).str.strip().str.casefold()
    target = market_name.strip().casefold()
    return df.loc[market_series == target].copy()

def find_next_url(soup: BeautifulSoup, current_url: str) -> str | None:
    # The ">" next button is an <a> whose visible text is ">"
    next_a = soup.find("a", string=lambda s: isinstance(s, str) and s.strip() == ">")
    if not next_a:
        return None

    href = next_a.get("href")
    if not href:
        return None

    # Make it absolute
    return urljoin(current_url, href)

def upsert_query_param(url: str, key: str, value: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[key] = value
    return urlunparse(parsed._replace(query=urlencode(query)))

def resolve_product_value(
    session: requests.Session,
    start_url: str,
    product_name: str,
    timeout: float,
    max_retries: int,
    retry_backoff_s: float,
) -> str:
    """Resolve product option value from the market page (e.g., Dry Maize -> 1)."""
    response = None
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            response = session.get(start_url, timeout=timeout)
            response.raise_for_status()
            break
        except requests.exceptions.RequestException as error:
            last_error = error
            if attempt < max_retries:
                time.sleep(retry_backoff_s * (2 ** (attempt - 1)))

    if response is None:
        error_msg = f"Unable to load market filter options after {max_retries} attempts. Last error: {last_error}"
        raise RuntimeError(error_msg) from last_error

    soup = BeautifulSoup(response.text, "lxml")
    select = soup.find("select", {"id": "selProduct"}) or soup.find("select", {"name": "product"})
    if not select:
        raise RuntimeError("Product filter select was not found on the market page.")

    target = product_name.strip().lower()
    for option in select.find_all("option"):
        label = option.get_text(" ", strip=True).lower()
        value = (option.get("value") or "").strip()
        if label == target and value:
            return value

    raise RuntimeError(f"Product filter option not found: {product_name}")

def scrape_all(
    start_url: str,
    sleep_s: float = 0.4,
    max_pages: int = 2000,
    verbose: bool = True,
    request_timeout: float = 40,
    max_retries: int = 3,
    retry_backoff_s: float = 2.0,
    product_filter_name: str = "Dry Maize",
    market_filter_name: str = "Eldoret Main",
    per_page: int = 3000,
    stop_event: Optional[threading.Event] = None,
    progress_callback=None,
) -> pd.DataFrame:
    session = requests.Session()
    session.headers.update(HEADERS)
    session.verify = False

    adapter = requests.adapters.HTTPAdapter(pool_connections=1, pool_maxsize=4)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    if progress_callback:
        progress_callback(status="Resolving product filter...", stage="Resolving", completed=0, total=None)

    try:
        filter_value = resolve_product_value(
            session=session,
            start_url=start_url,
            product_name=product_filter_name,
            timeout=request_timeout,
            max_retries=max_retries,
            retry_backoff_s=retry_backoff_s,
        )

        url = upsert_query_param(start_url, "product", filter_value)
        url = upsert_query_param(url, "per_page", str(per_page))

        all_dfs: list[pd.DataFrame] = []
        seen: set[str] = set()
        pages_with_data = 0
        last_columns = None
        discovered_total_pages: Optional[int] = None
        eta = EtaEstimator()

        for page_idx in range(max_pages):
            page_no = page_idx + 1

            if stop_event and stop_event.is_set():
                raise RuntimeError("Scrape stopped by user.")

            if url in seen:
                break
            seen.add(url)

            if progress_callback:
                progress_callback(
                    status=f"Fetching page {page_no}...",
                    stage="Fetching",
                    completed=page_no - 1,
                    total=discovered_total_pages,
                    current_item=page_no,
                    eta_seconds=eta.eta_seconds(page_no - 1, discovered_total_pages),
                )

            r = None
            last_error = None
            for attempt in range(1, max_retries + 1):
                if stop_event and stop_event.is_set():
                    raise RuntimeError("Scrape stopped by user.")
                try:
                    r = session.get(url, timeout=request_timeout)
                    r.raise_for_status()
                    break
                except requests.exceptions.RequestException as error:
                    last_error = error
                    force_wait_s = 0.0
                    if isinstance(error, requests.exceptions.HTTPError) and error.response is not None:
                        if error.response.status_code == 429:
                            force_wait_s = float(error.response.headers.get("Retry-After", 60))
                    if attempt < max_retries:
                        wait_s = force_wait_s if force_wait_s else retry_backoff_s * (2 ** (attempt - 1))
                        time.sleep(wait_s)

            if r is None:
                if all_dfs:
                    break
                raise RuntimeError("Unable to fetch the first page.") from last_error

            page_soup = BeautifulSoup(r.text, "lxml")
            discovered = estimate_total_pages_from_soup(page_soup)
            if discovered:
                discovered_total_pages = discovered if discovered_total_pages is None else max(discovered_total_pages, discovered)

            df = extract_table(r.text)
            if is_effectively_empty_table(df):
                break

            pages_with_data += 1
            last_columns = list(df.columns)

            market_df = filter_market_rows(df, market_filter_name)
            if not market_df.empty:
                all_dfs.append(market_df)

            eta.mark_item_complete()
            if progress_callback:
                progress_callback(
                    status=f"Processed page {page_no}",
                    stage="Parsing",
                    completed=page_no,
                    total=discovered_total_pages,
                    current_item=page_no,
                    eta_seconds=eta.eta_seconds(page_no, discovered_total_pages),
                )

            nxt = find_next_url(page_soup, url)
            if not nxt:
                break

            url = nxt
            jitter_s = sleep_s + random.uniform(0.0, sleep_s * 0.5)
            if stop_event and stop_event.is_set():
                raise RuntimeError("Scrape stopped by user.")
            time.sleep(jitter_s)

        if not all_dfs:
            if pages_with_data == 0:
                raise RuntimeError("No table data was scraped.")
            return pd.DataFrame(columns=last_columns or [])

        out = pd.concat(all_dfs, ignore_index=True)
        dedup_cols = [c for c in out.columns if c in {"Date", "Market", "Commodity", "Wholesale Price", "Retail Price"}]
        if dedup_cols:
            out = out.drop_duplicates(subset=dedup_cols)

        return out
    finally:
        session.close()


def save_to_excel_with_table(df: pd.DataFrame, output_path: str) -> None:
    """Save dataframe to Excel and apply Orange 'Table Style Medium 7' table styling."""
    df.to_excel(output_path, index=False, sheet_name="Data")

    # Apply native Excel table styling via openpyxl.
    wb = load_workbook(output_path)
    ws = wb["Data"]

    if ws.max_column == 0:
        wb.save(output_path)
        return

    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row

    # Excel table requires at least a header and one data row.
    if last_row < 2:
        wb.save(output_path)
        return

    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName="KamisData", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Autofit column widths based on the longest cell content in each column.
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            max_len = max(max_len, len(str(cell_value)))

        # Add small padding and cap width to keep sheets readable.
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


class OnionPriceCheckerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Onion Price Checker v1.00")
        self.root.geometry("920x500")
        self.root.minsize(860, 460)

        self.msg_queue: queue.Queue[dict] = queue.Queue()
        self.worker_thread: Optional[threading.Thread] = None
        self.stop_event = threading.Event()
        self.pending_close = False
        self.run_started_at = 0.0
        self.last_output_uri: Optional[str] = None
        self.spinner_frames = ["|", "/", "-", "\\"]
        self.spinner_index = 0
        self.spinner_job: Optional[str] = None

        self.state_var = tk.StringVar(value="Idle")
        self.status_var = tk.StringVar(value="Ready")
        self.items_var = tk.StringVar(value="0 / ?")
        self.spinner_var = tk.StringVar(value="-")

        self._build_ui()
        self._set_idle_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.after(150, self.process_queue)

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=14)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Onion Price Checker v1.00", font=("Segoe UI", 15, "bold")).pack(anchor="w")

        button_row = ttk.Frame(main)
        button_row.pack(fill="x", pady=(12, 10))

        self.start_btn = ttk.Button(button_row, text="Start", command=self.start_scrape)
        self.start_btn.pack(side="left")

        self.stop_btn = ttk.Button(button_row, text="Stop", command=self.stop_scrape)
        self.stop_btn.pack(side="left", padx=(8, 0))

        ttk.Frame(button_row).pack(side="left", fill="x", expand=True)
        self.close_btn = ttk.Button(button_row, text="Close", command=self.close_app)
        self.close_btn.pack(side="right")

        status_frame = ttk.LabelFrame(main, text="Status", padding=10)
        status_frame.pack(fill="x")

        ttk.Label(status_frame, textvariable=self.state_var).pack(anchor="w")
        ttk.Label(status_frame, textvariable=self.status_var).pack(anchor="w", pady=(4, 0))

        spinner_row = ttk.Frame(main)
        spinner_row.pack(fill="x", pady=(14, 8))
        ttk.Label(spinner_row, text="Activity:").pack(side="left")
        ttk.Label(spinner_row, textvariable=self.spinner_var, font=("Consolas", 16, "bold")).pack(side="left", padx=(8, 0))

        metrics_row = ttk.Frame(main)
        metrics_row.pack(fill="x")

        ttk.Label(metrics_row, text="Items:").grid(row=0, column=0, sticky="w")
        ttk.Label(metrics_row, textvariable=self.items_var).grid(row=0, column=1, sticky="w", padx=(6, 0))

        result_frame = ttk.LabelFrame(main, text="Result", padding=10)
        result_frame.pack(fill="both", expand=True, pady=(12, 0))

        self.result_text = tk.Text(result_frame, height=8, wrap="none")
        self.result_text.pack(side="left", fill="both", expand=True)
        yscroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_text.yview)
        yscroll.pack(side="right", fill="y")
        self.result_text.configure(yscrollcommand=yscroll.set)
        self.result_text.configure(state="disabled")

        link_row = ttk.Frame(main)
        link_row.pack(fill="x", pady=(10, 0))

        ttk.Label(link_row, text="Produced file:").pack(side="left")
        self.open_link = tk.Label(
            link_row,
            text="No file yet",
            fg="gray",
            cursor="arrow",
            anchor="w",
        )
        self.open_link.pack(side="left", padx=(8, 0), fill="x", expand=True)
        self.open_link.bind("<Button-1>", self.on_open_link_clicked)

    def set_result_text(self, text: str) -> None:
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", tk.END)
        self.result_text.insert(tk.END, text)
        self.result_text.configure(state="disabled")

    def _set_idle_ui(self) -> None:
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.stop_spinner()

    def tick_spinner(self) -> None:
        self.spinner_var.set(self.spinner_frames[self.spinner_index])
        self.spinner_index = (self.spinner_index + 1) % len(self.spinner_frames)
        self.spinner_job = self.root.after(120, self.tick_spinner)

    def start_spinner(self) -> None:
        if self.spinner_job is None:
            self.spinner_index = 0
            self.tick_spinner()

    def stop_spinner(self) -> None:
        if self.spinner_job is not None:
            self.root.after_cancel(self.spinner_job)
            self.spinner_job = None
        self.spinner_var.set("-")

    def set_open_file_link(self, file_uri: Optional[str], text: str) -> None:
        self.last_output_uri = file_uri
        if file_uri:
            self.open_link.config(text=text, fg="#005A9C", cursor="hand2")
        else:
            self.open_link.config(text=text, fg="gray", cursor="arrow")

    def on_open_link_clicked(self, _event=None) -> None:
        if not self.last_output_uri:
            return
        try:
            webbrowser.open(self.last_output_uri)
        except Exception as error:
            messagebox.showerror("Open file", f"Could not open file link.\n\n{error}")

    def close_app(self) -> None:
        """Close button handler; delegates to safe window-close behavior."""
        self.on_close()

    def start_scrape(self) -> None:
        if self.worker_thread and self.worker_thread.is_alive():
            return

        self.stop_event.clear()
        self.pending_close = False
        self.run_started_at = time.perf_counter()
        self.set_result_text("")
        self.set_open_file_link(None, "No file yet")
        self.state_var.set("Running")
        self.status_var.set("Starting scraper...")
        self.items_var.set("0 / ?")

        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.start_spinner()

        self.worker_thread = threading.Thread(target=self.worker_run_scraper, daemon=True)
        self.worker_thread.start()

    def stop_scrape(self) -> None:
        if not self.worker_thread or not self.worker_thread.is_alive():
            return
        self.state_var.set("Stopping")
        self.status_var.set("Waiting for safe stop...")
        self.stop_event.set()
        self.stop_btn.config(state="disabled")

    def worker_run_scraper(self) -> None:
        def emit(**payload) -> None:
            self.msg_queue.put(payload)

        try:
            df = scrape_all(
                START_URL,
                verbose=False,
                product_filter_name=TARGET_PRODUCT,
                market_filter_name=TARGET_MARKET,
                per_page=TARGET_PER_PAGE,
                stop_event=self.stop_event,
                progress_callback=emit,
            )

            if self.stop_event.is_set():
                self.msg_queue.put({"kind": "stopped"})
                return

            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
                current_year = datetime.now().year
                valid_years = [current_year, current_year - 1]
                valid_date_mask = df["Date"].notna()
                year_mask = df[valid_date_mask]["Date"].dt.year.isin(valid_years)
                df = df[valid_date_mask][year_mask]

            df = df.drop(columns=["Grade", "Sex"], errors="ignore")

            now = datetime.now()
            timestamp = f"{now.month}-{now.day}-{now.year}  {now.strftime('%I.%M.%S %p')}"
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_file = OUTPUT_DIR / f"onion_price_checker_v1.00_{timestamp}.xlsx"
            save_to_excel_with_table(df, str(output_file))

            self.msg_queue.put({
                "kind": "completed",
                "rows": len(df),
                "file_link": output_file.resolve().as_uri(),
                "file_path": str(output_file),
                "elapsed": time.perf_counter() - self.run_started_at,
            })
        except Exception as error:
            if self.stop_event.is_set():
                self.msg_queue.put({"kind": "stopped"})
            else:
                self.msg_queue.put({"kind": "error", "message": str(error)})

    def process_queue(self) -> None:
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                kind = msg.get("kind")

                if kind == "completed":
                    self.state_var.set("Completed")
                    self.status_var.set("Scrape completed successfully")
                    self.set_open_file_link(msg["file_link"], msg["file_path"])
                    self.set_result_text(
                        f"Rows: {msg['rows']:,}\n"
                        f"File: {msg['file_path']}\n"
                        f"Open file: {msg['file_link']}\n"
                        f"Elapsed: {msg['elapsed']:.2f}s"
                    )
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                elif kind == "stopped":
                    self.state_var.set("Stopped")
                    self.status_var.set("Scrape stopped safely")
                    self.set_open_file_link(None, "No file produced")
                    self.set_result_text("No output file was produced because the run was stopped.")
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                elif kind == "error":
                    self.state_var.set("Idle")
                    self.status_var.set("Failed")
                    self.set_open_file_link(None, "No file produced")
                    self.set_result_text(f"Error: {msg.get('message', 'Unknown error')}")
                    self._set_idle_ui()
                    if self.pending_close:
                        self.root.destroy()

                else:
                    completed = int(msg.get("completed", 0))
                    total = msg.get("total")
                    self.status_var.set(msg.get("status", "Running..."))

                    if total and total > 0:
                        self.items_var.set(f"{completed} / {total}")
                    else:
                        self.items_var.set(f"{completed} / ?")

        except queue.Empty:
            pass
        finally:
            self.root.after(150, self.process_queue)

    def on_close(self) -> None:
        running = self.worker_thread and self.worker_thread.is_alive()
        if running:
            should_stop = messagebox.askyesno(
                "Stop scraper?",
                "A scrape is in progress. Stop safely and close the app?",
            )
            if not should_stop:
                return
            self.pending_close = True
            self.stop_scrape()
            self.state_var.set("Stopping")
            self.status_var.set("Stopping before close...")
            return
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = OnionPriceCheckerApp(root)
    root.mainloop()
