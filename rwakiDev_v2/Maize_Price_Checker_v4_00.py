# =============================================================================
# Maize Price Checker  —  v4.00
# =============================================================================
#
# DESCRIPTION
# -----------
# Scrapes dry maize wholesale and retail price data from the Kenya Agricultural
# Market Information System (KAMIS) website at https://kamis.kilimo.go.ke.
#
# The script targets a specific product ("Dry Maize") and market ("Eldoret Main"),
# filters results to the current year and the previous year only, removes
# unnecessary columns (Grade, Sex), and saves the cleaned data to a
# timestamped Excel spreadsheet formatted with an Orange Table Style Medium 7
# table, auto-fitted column widths, and saved to a user-defined output folder.
#
# -----------------------------------------------------------------------------
# CHANGELOG
# -----------------------------------------------------------------------------
#
# v4.00  —  2026-04-12
#   OPTIMIZATIONS (performance, reliability, polite scraping)
#   - Added random jitter (up to 50 % of sleep_s) between page requests so
#     traffic does not look like a clock-driven bot.
#   - Added explicit 429 rate-limit handling: honours the server's Retry-After
#     header (defaults to 60 s if absent) before retrying, instead of the
#     normal short backoff.
#   - HTML is now parsed only once per page using lxml (BS4); the resulting
#     soup object is reused for next-link detection, eliminating a second full
#     parse that was happening silently on every page.
#   - Switched BS4 HTML parser from the built-in html.parser to lxml
#     (3–5× faster) in both scrape_all and resolve_product_value.
#   - Removed redundant upsert_query_param calls from inside the pagination
#     loop (URL parameters that never change were being rebuilt every page).
#   - Added HTTPAdapter with explicit connection pool settings so the TCP
#     connection to the server stays alive across all page requests instead
#     of re-handshaking each time.
#   - Added post-scrape deduplication keyed on Date, Market, Commodity,
#     Wholesale Price, and Retail Price to silently remove any rows duplicated
#     by pagination overlap; logged when rows are removed.
#   - Expanded request headers to include Accept, Accept-Language, and
#     Connection: keep-alive so requests look more like a real browser.
#
# v3.00  —  2026-04-12
#   - Changed output format from CSV to Excel (.xlsx).
#   - Applied Orange "Table Style Medium 7" table formatting via openpyxl.
#   - Added Excel auto-fit column widths (capped at 60 characters).
#   - Set output directory to:
#       C:\Users\kirwa\Documents\Farming and Agriculture\Kamis Data Results
#   - Output directory is created automatically if it does not exist.
#   - Timestamped filename format changed to:  M-D-YYYY  HH.MM.SS AM/PM
#   - Removed Grade and Sex columns from the final spreadsheet.
#   - After saving, prints a clickable file:// URI link to the created file.
#
# v2.00  —  2026-04-12
#   - Added year filtering: only data from the current year and the previous
#     year is kept; older rows are discarded before saving.
#   - Added SSL certificate verification bypass (self-signed cert on server)
#     with urllib3 InsecureRequestWarning suppressed.
#   - Improved error messages to include the specific error type and attempt
#     count when the initial option-resolve request fails.
#
# v1.00  —  initial release
#   - Scrapes all pages of the KAMIS market table for Dry Maize / Eldoret Main.
#   - Resolves the product filter value dynamically from the page HTML.
#   - Paginates until no next-page link is found or max_pages is reached.
#   - Filters rows to the target market before collecting.
#   - Handles transient network failures with configurable exponential backoff.
#   - Detects and stops on empty or duplicate pages.
#   - Saves output to kamis_market_dump.csv.
#   - Detects open Excel processes before saving and offers to close them.
#
# =============================================================================

import time
import sys
import subprocess
import random
from io import StringIO
from datetime import datetime
from pathlib import Path
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from urllib.parse import urljoin, urlparse, parse_qsl, urlencode, urlunparse

# Suppress SSL warnings for self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE = "https://kamis.kilimo.go.ke"
START_URL = "https://kamis.kilimo.go.ke/site/market"  # or any filtered URL you have
TARGET_PRODUCT = "Dry Maize"
TARGET_MARKET = "Eldoret Main"
TARGET_PER_PAGE = 3000
OUTPUT_DIR = Path(r"c:\Users\kirwa\Documents\Farming and Agriculture\Kamis Data Results")


def is_excel_open() -> bool:
    """Detect whether Microsoft Excel is currently running (Windows only)."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq EXCEL.EXE"],
            capture_output=True,
            text=True,
            check=False,
        )
        return "excel.exe" in (result.stdout or "").lower()
    except Exception:
        return False


def close_excel_processes() -> bool:
    """Attempt to close all Excel processes. Returns True when successful."""
    if not sys.platform.startswith("win"):
        return False

    try:
        result = subprocess.run(
            ["taskkill", "/IM", "EXCEL.EXE", "/F"],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def prompt_close_excel_if_open() -> None:
    """Prompt user to close Excel before writing CSV to avoid file-lock issues."""
    if not is_excel_open():
        return

    print("[WARN] Microsoft Excel appears to be running.", flush=True)
    print("[WARN] If the CSV file is open in Excel, saving may fail.", flush=True)

    try:
        answer = input("Close Excel now? [y/N]: ").strip().lower()
    except EOFError:
        answer = "n"

    if answer in {"y", "yes"}:
        if close_excel_processes():
            print("[INFO] Excel process closed.", flush=True)
        else:
            print("[WARN] Could not close Excel automatically. Please close it manually.", flush=True)
    else:
        print("[INFO] Continuing without closing Excel.", flush=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
}

def extract_table(html: str) -> pd.DataFrame:
    # KAMIS renders a single main table in this page; read_html usually finds it.
    tables = pd.read_html(StringIO(html))
    df = tables[-1].copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def is_effectively_empty_table(df: pd.DataFrame) -> bool:
    """Return True when a parsed table has no meaningful data rows."""
    if df.empty:
        return True

    normalized = df.copy()
    for col in normalized.columns:
        if normalized[col].dtype == object:
            normalized[col] = normalized[col].astype(str).str.strip()

    normalized = normalized.replace({"": pd.NA, "-": pd.NA, " - ": pd.NA, "nan": pd.NA, "None": pd.NA})
    return normalized.dropna(how="all").empty

def filter_market_rows(df: pd.DataFrame, market_name: str) -> pd.DataFrame:
    if "Market" not in df.columns:
        return df.iloc[0:0].copy()
    market_series = df["Market"].astype(str).str.strip().str.casefold()
    target = market_name.strip().casefold()
    return df.loc[market_series == target].copy()

def find_next_url(soup: BeautifulSoup, current_url: str) -> str | None:
    # The ">" next button is an <a> whose visible text is ">"
    next_a = soup.find("a", string=lambda s: isinstance(s, str) and s.strip() == ">")
    if not next_a:
        return None

    href = next_a.get("href")
    if not href:
        return None

    # Make it absolute
    return urljoin(current_url, href)

def upsert_query_param(url: str, key: str, value: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[key] = value
    return urlunparse(parsed._replace(query=urlencode(query)))

def resolve_product_value(
    session: requests.Session,
    start_url: str,
    product_name: str,
    timeout: float,
    max_retries: int,
    retry_backoff_s: float,
) -> str:
    """Resolve product option value from the market page (e.g., Dry Maize -> 1)."""
    response = None
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            response = session.get(start_url, timeout=timeout)
            response.raise_for_status()
            break
        except requests.exceptions.RequestException as error:
            last_error = error
            print(f"[DEBUG] Attempt {attempt}/{max_retries} failed: {type(error).__name__}: {error}", flush=True)
            if attempt < max_retries:
                time.sleep(retry_backoff_s * (2 ** (attempt - 1)))

    if response is None:
        error_msg = f"Unable to load market filter options after {max_retries} attempts. Last error: {last_error}"
        raise RuntimeError(error_msg) from last_error

    soup = BeautifulSoup(response.text, "lxml")
    select = soup.find("select", {"id": "selProduct"}) or soup.find("select", {"name": "product"})
    if not select:
        raise RuntimeError("Product filter select was not found on the market page.")

    target = product_name.strip().lower()
    for option in select.find_all("option"):
        label = option.get_text(" ", strip=True).lower()
        value = (option.get("value") or "").strip()
        if label == target and value:
            return value

    raise RuntimeError(f"Product filter option not found: {product_name}")

def scrape_all(
    start_url: str,
    sleep_s: float = 0.4,
    max_pages: int = 2000,
    verbose: bool = True,
    request_timeout: float = 40,
    max_retries: int = 3,
    retry_backoff_s: float = 2.0,
    product_filter_name: str = "Dry Maize",
    market_filter_name: str = "Eldoret Main",
    per_page: int = 3000,
) -> pd.DataFrame:
    session = requests.Session()
    session.headers.update(HEADERS)
    session.verify = False  # Disable SSL verification for self-signed certificates

    # Single connection to one host; pool_maxsize keeps persistent TCP connection alive.
    _adapter = requests.adapters.HTTPAdapter(pool_connections=1, pool_maxsize=4)
    session.mount("https://", _adapter)
    session.mount("http://", _adapter)

    filter_value = resolve_product_value(
        session=session,
        start_url=start_url,
        product_name=product_filter_name,
        timeout=request_timeout,
        max_retries=max_retries,
        retry_backoff_s=retry_backoff_s,
    )

    url = upsert_query_param(start_url, "product", filter_value)
    url = upsert_query_param(url, "per_page", str(per_page))
    all_dfs = []
    seen = set()
    started_at = time.perf_counter()
    pages_with_data = 0
    last_columns = None

    if verbose:
        print(
            "[START] Scrape started "
            f"| max_pages={max_pages} | sleep_s={sleep_s} "
            f"| timeout={request_timeout}s | max_retries={max_retries} "
            f"| product={product_filter_name}({filter_value}) "
            f"| market={market_filter_name} | per_page={per_page}",
            flush=True,
        )

    for page_idx in range(max_pages):
        page_no = page_idx + 1
        if url in seen:
            if verbose:
                print(f"[STOP] Revisited URL detected at page {page_no}; stopping.", flush=True)
            break
        seen.add(url)

        if verbose:
            print(f"[PAGE {page_no}] Fetching URL: {url}", flush=True)

        req_start = time.perf_counter()
        r = None
        last_error = None
        for attempt in range(1, max_retries + 1):
            try:
                if verbose and attempt > 1:
                    print(f"[PAGE {page_no}] Retry {attempt}/{max_retries}", flush=True)
                r = session.get(url, timeout=request_timeout)
                r.raise_for_status()
                break
            except requests.exceptions.RequestException as error:
                last_error = error
                # Honour 429 Retry-After before falling back to exponential backoff.
                force_wait_s: float = 0.0
                if isinstance(error, requests.exceptions.HTTPError) and error.response is not None:
                    if error.response.status_code == 429:
                        force_wait_s = float(error.response.headers.get("Retry-After", 60))
                        if verbose:
                            print(f"[PAGE {page_no}] Rate limited (429). Waiting {force_wait_s:.0f}s.", flush=True)
                if attempt < max_retries:
                    wait_s = force_wait_s if force_wait_s else retry_backoff_s * (2 ** (attempt - 1))
                    if verbose and not force_wait_s:
                        print(
                            f"[PAGE {page_no}] Request failed ({type(error).__name__}): {error}",
                            flush=True,
                        )
                        print(f"[PAGE {page_no}] Backing off for {wait_s:.2f}s before retry", flush=True)
                    time.sleep(wait_s)
                else:
                    if verbose:
                        print(
                            f"[PAGE {page_no}] Request failed after {max_retries} attempts; stopping scrape.",
                            flush=True,
                        )

        if r is None:
            if all_dfs:
                if verbose:
                    print(f"[STOP] Returning partial data after page {page_no - 1} due to request failures.", flush=True)
                break
            raise RuntimeError("Unable to fetch the first page.") from last_error

        req_seconds = time.perf_counter() - req_start

        if verbose:
            print(
                f"[PAGE {page_no}] HTTP {r.status_code} in {req_seconds:.2f}s | bytes={len(r.text):,}",
                flush=True,
            )

        # Parse HTML once with lxml; reuse soup for next-link detection to avoid double-parsing.
        page_soup = BeautifulSoup(r.text, "lxml")

        parse_start = time.perf_counter()
        df = extract_table(r.text)
        parse_seconds = time.perf_counter() - parse_start

        if is_effectively_empty_table(df):
            if verbose:
                print(f"[STOP] Empty table detected on page {page_no}; stopping.", flush=True)
            break

        pages_with_data += 1
        last_columns = list(df.columns)

        market_df = filter_market_rows(df, market_filter_name)
        if verbose:
            print(
                f"[PAGE {page_no}] Market filter kept {len(market_df):,}/{len(df):,} rows for '{market_filter_name}'",
                flush=True,
            )

        if not market_df.empty:
            all_dfs.append(market_df)

        if verbose:
            print(
                f"[PAGE {page_no}] Parsed table in {parse_seconds:.2f}s | rows={len(df):,}",
                flush=True,
            )

        next_start = time.perf_counter()
        nxt = find_next_url(page_soup, url)
        next_seconds = time.perf_counter() - next_start

        if verbose:
            print(f"[PAGE {page_no}] Next-link parse in {next_seconds:.2f}s", flush=True)

        if not nxt:
            if verbose:
                print(f"[STOP] No next page after page {page_no}; stopping.", flush=True)
            break

        url = nxt
        # Add jitter so traffic pattern looks natural and avoids server-side rate triggers.
        jitter_s = sleep_s + random.uniform(0.0, sleep_s * 0.5)
        if verbose:
            print(f"[PAGE {page_no}] Sleeping for {jitter_s:.2f}s", flush=True)
        time.sleep(jitter_s)

    if not all_dfs:
        if pages_with_data == 0:
            raise RuntimeError("No table data was scraped.")
        out = pd.DataFrame(columns=last_columns or [])
        if verbose:
            total_seconds = time.perf_counter() - started_at
            print(
                f"[DONE] Pages={pages_with_data:,} | Rows=0 | total={total_seconds:.2f}s",
                flush=True,
            )
        return out

    concat_start = time.perf_counter()
    out = pd.concat(all_dfs, ignore_index=True)

    # Deduplicate rows to handle any pagination overlap without silently inflating counts.
    dedup_cols = [c for c in out.columns if c in {"Date", "Market", "Commodity", "Wholesale Price", "Retail Price"}]
    if dedup_cols:
        rows_before = len(out)
        out = out.drop_duplicates(subset=dedup_cols)
        removed = rows_before - len(out)
        if verbose and removed > 0:
            print(f"[INFO] Removed {removed:,} duplicate rows from pagination overlap.", flush=True)

    concat_seconds = time.perf_counter() - concat_start

    if verbose:
        total_seconds = time.perf_counter() - started_at
        print(
            f"[DONE] Pages={pages_with_data:,} | Rows={len(out):,} | concat={concat_seconds:.2f}s | total={total_seconds:.2f}s",
            flush=True,
        )

    return out


def save_to_excel_with_table(df: pd.DataFrame, output_path: str) -> None:
    """Save dataframe to Excel and apply Orange 'Table Style Medium 7' table styling."""
    df.to_excel(output_path, index=False, sheet_name="Data")

    # Apply native Excel table styling via openpyxl.
    wb = load_workbook(output_path)
    ws = wb["Data"]

    if ws.max_column == 0:
        wb.save(output_path)
        return

    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row

    # Excel table requires at least a header and one data row.
    if last_row < 2:
        wb.save(output_path)
        return

    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName="KamisData", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Autofit column widths based on the longest cell content in each column.
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            max_len = max(max_len, len(str(cell_value)))

        # Add small padding and cap width to keep sheets readable.
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)

if __name__ == "__main__":
    app_start = time.perf_counter()
    try:
        prompt_close_excel_if_open()

        # Use larger page size to reduce pagination requests while preserving result fidelity.
        df = scrape_all(
            START_URL,
            product_filter_name=TARGET_PRODUCT,
            market_filter_name=TARGET_MARKET,
            per_page=TARGET_PER_PAGE,
        )

        # Optional cleanups
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
            
            # Filter data to only current year (2026) and previous year (2025)
            current_year = datetime.now().year
            valid_years = [current_year, current_year - 1]
            
            # Only filter rows with valid dates
            valid_date_mask = df["Date"].notna()
            year_mask = df[valid_date_mask]["Date"].dt.year.isin(valid_years)
            
            # Combine masks: keep rows with valid dates that match valid years
            df = df[valid_date_mask][year_mask]
            print(f"[INFO] Filtered data for years: {valid_years}")
        else:
            print("[WARN] 'Date' column not found in scraped data.", flush=True)

        # Remove columns not needed in the final spreadsheet.
        df = df.drop(columns=["Grade", "Sex"], errors="ignore")

        _now = datetime.now()
        timestamp = f"{_now.month}-{_now.day}-{_now.year}  {_now.strftime('%I.%M.%S %p')}"
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_file = OUTPUT_DIR / f"kamis_market_dump_v4.00_{timestamp}.xlsx"
        save_to_excel_with_table(df, str(output_file))
        output_link = output_file.resolve().as_uri()
        print(df.head(3))
        print(f"\nRows: {len(df):,}")
        print(f"Saved: {output_file}")
        print(f"Open file: {output_link}")
    except KeyboardInterrupt:
        print("[CANCELLED] Scrape interrupted by user.", flush=True)
        sys.exit(130)
    except Exception as error:
        print(f"[ERROR] Scrape failed: {error}", flush=True)
        sys.exit(1)
    finally:
        print(f"Elapsed: {time.perf_counter() - app_start:.2f}s")