"""
Speed Test GUI Application v1.08
================================================================================
CHANGELOG - Version 1.08
================================================================================
V1.08 CHANGES:
1. Added a new indigo "Multi-Test" button next to the Re-test button.
2. Multi-Test runs three providers back-to-back so a single slow result can be
   distinguished from a slow line:
       - Cloudflare      (reuses existing v1.06 measurement code)
       - Netflix Fast.com (parallel CDN streams via api.fast.com)
       - Ookla speedtest CLI (only if `speedtest` binary is on PATH)
3. Added a MULTI-PROVIDER COMPARISON table that shows download, upload, ping,
   jitter, and the test server for each provider in one place.
4. Added spread / ratio analysis: if the fastest provider is more than 2x the
   slowest, the report calls out which provider is the bottleneck and points
   at a likely peering/routing issue rather than the user's line.
5. Added an append-only multi_provider_log.csv so multi-tests are tracked
   separately from the single-provider speed log.
6. Ookla provider is optional — if the CLI is not installed the row is
   skipped with a hint pointing at https://www.speedtest.net/apps/cli.
7. Stop button now interrupts in-progress Fast.com streams and terminates
   the Ookla subprocess.

V1.07 CHANGES:
1. Added a dark-red "Error Log" button for manual diagnostic runs.
2. Added automatic diagnostics when download is below 15 Mbps or upload is below 5 Mbps.
3. Added DNS, TCP, TLS, HTTP, proxy, VPN, and network-interface probes.
4. Added Norton, NordVPN, Zscaler, and other inspecting-middlebox detection.
5. Added append-only output to `internet_speed_diagnostics.log`.

V1.06 CHANGES:
1. Duplicated v1.05 into a new graph-enabled version.
2. Added a "View Graph" button that reads `internet_speed_log.csv`.
3. Creates a professional dark-mode futuristic line graph in a new window.
4. Splits telemetry into two stacked charts: internet speed on top and
   ping/jitter latency on the bottom.
5. Plots every valid row from the speed log in chronological order.
6. Uses pure Tkinter Canvas drawing so the graph works without extra packages.

V1.05 CHANGES:
1. Duplicated v1.04 into a new Cloudflare-backed version.
2. Replaced the `speedtest` library with direct calls to Cloudflare Speed Test endpoints.
3. Uses Cloudflare `/meta` for connection/IP/ISP details, `__down` for download,
   `__up` for upload, and lightweight requests for latency/jitter.
4. Keeps the existing Tkinter interface and CSV logging format.

V1.04 CHANGES:
1. Enhanced error handling for all HTTP error codes (not just 403)
2. Added extract_error_code() method to parse HTTP status codes from exceptions
3. Improved error reporting to include detailed server information.
4. Server details now displayed when ANY error occurs (not just 403)
5. Better user feedback during server retry operations

PREVIOUS (V1.03) CHANGES:
1. FIRST RELEASE BUILD - Production-ready version
2. Packaged as standalone executable with digital signature
3. Includes comprehensive speed testing with 403 error retry logic
4. Full CSV logging and comment support for test results

PREVIOUS (V1.02) FEATURES:
1. Added automatic retry mechanism for 403 Forbidden errors from servers
2. Retries up to 3 times to find an alternate working server
3. Logs all 403 errors with server sponsor and location details

PREVIOUS (V1.01) FEATURES:
1. Added purple "Log" button to save test results to CSV file
2. Implemented CSV logging to 'internet_speed_log.csv' with full data export
3. Newest test results are prepended to CSV (top row always latest)
4. Added comment input field; log button disabled until results exist

WHAT THIS CODE DOES
1. Provides a desktop Tkinter GUI to run internet speed tests using Cloudflare,
   Netflix Fast.com, and (optionally) Ookla CLI.
2. Displays real-time progress and detailed output in a scrollable text panel.
3. Collects and shows key metrics: download speed, upload speed, ping, and jitter.
4. Captures Cloudflare edge and network details (colo, location, country, coordinates, ISP, and IP).
5. Handles HTTP/network errors with clearer diagnostics.
6. Allows users to add short comments and log test results to `internet_speed_log.csv`.
7. Writes CSV output with the newest result at the top so recent tests are easy to find.
8. Opens a built-in graph view for visual trend analysis of the speed log.
9. Multi-provider mode triangulates ISP vs path-specific bottlenecks.

================================================================================
"""

import tkinter as tk
from tkinter import messagebox, scrolledtext
import threading
from datetime import datetime
import csv
import json
import ipaddress
import os
import platform
import queue
import shutil
import socket
import ssl
import subprocess
import sys
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Table = None
    TableStyleInfo = None
    get_column_letter = None

# Fix for PyInstaller GUI mode - redirect stdout/stderr
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')


class SpeedTestApp:
    EXPECTED_MIN_DOWNLOAD_MBPS = 15.0
    EXPECTED_MIN_UPLOAD_MBPS = 5.0
    DIAGNOSTIC_HOSTS = [
        ("speed.cloudflare.com", 443),
        ("1.1.1.1", 443),
        ("8.8.8.8", 443),
        ("www.google.com", 443),
    ]
    CLOUDFLARE_BASE_URL = "https://speed.cloudflare.com"
    UPLOAD_SAMPLE_SIZES = [500_000, 1_000_000, 2_500_000, 5_000_000]
    UPLOAD_RETRIES = 2

    # Multi-provider configuration (v1.08)
    FASTDOTCOM_TOKEN = "YXNkZmFzZGxmbnNkYWZoYXNkZmhrYWxm"
    FASTDOTCOM_API = (
        "https://api.fast.com/netflix/speedtest/v2"
        "?https=true&token={token}&urlCount={count}"
    )
    FAST_PARALLEL_STREAMS = 3
    FAST_SAMPLE_BYTES = 25_000_000
    FAST_STREAM_TIMEOUT = 30
    OOKLA_TIMEOUT_SECONDS = 120
    MULTI_PROVIDER_LOG = "multi_provider_log.csv"

    DARK_THEME = {
        "window": "#12161f",
        "panel": "#1a2030",
        "field": "#0f131b",
        "text": "#e8edf7",
        "muted": "#a9b4c7",
        "border": "#2a3346",
        "select_bg": "#2f6fed",
        "select_fg": "#ffffff",
        "disabled": "#7f8797",
    }
    GRAPH_COLORS = {
        "background": "#090d14",
        "panel": "#0d1420",
        "grid": "#243247",
        "axis": "#5b6d89",
        "text": "#e6edf7",
        "muted": "#8ea0ba",
        "download": "#18f2b2",
        "upload": "#ffd166",
        "ping": "#4da3ff",
        "jitter": "#ff4fd8",
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Speed Test Application v1.08 - Cloudflare + Fast.com + Ookla")
        self.root.geometry("900x620")

        # Output text area
        self.output_frame = tk.Frame(root)
        self.output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.output_text = scrolledtext.ScrolledText(
            self.output_frame,
            wrap=tk.WORD,
            font=("Courier", 10),
            height=28,
            width=100
        )
        self.output_text.pack(fill=tk.BOTH, expand=True)
        self.output_text.insert(tk.END, "Click 'Start Test' for a quick Cloudflare test, or 'Multi-Test' to compare providers...\n")

        # Comment entry (above buttons)
        self.comment_frame = tk.Frame(root)
        self.comment_frame.pack(fill=tk.X, padx=10, pady=(0, 6))
        self.comment_label = tk.Label(self.comment_frame, text="Comments, Keep it Short", font=("Arial", 10))
        self.comment_label.pack(anchor='w')
        self.comment_entry = tk.Entry(self.comment_frame)
        self.comment_entry.pack(fill=tk.X, expand=True, padx=0, pady=(4, 0))

        # Button frame
        self.button_frame = tk.Frame(root)
        self.button_frame.pack(fill=tk.X, padx=10, pady=6)

        # Start Test button
        self.start_button = tk.Button(
            self.button_frame,
            text="Start Test",
            command=self.start_test,
            width=15,
            height=2,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.start_button.pack(side=tk.LEFT, padx=5)

        # Re-test button
        self.retest_button = tk.Button(
            self.button_frame,
            text="Re-test",
            command=self.start_test,
            width=15,
            height=2,
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.retest_button.pack(side=tk.LEFT, padx=5)

        # Multi-Test button (v1.08)
        self.multi_test_button = tk.Button(
            self.button_frame,
            text="Multi-Test",
            command=self.start_multi_test,
            width=15,
            height=2,
            bg="#3F51B5",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.multi_test_button.pack(side=tk.LEFT, padx=5)

        # Stop button
        self.stop_button = tk.Button(
            self.button_frame,
            text="Stop",
            command=self.stop_test,
            width=15,
            height=2,
            bg="#FF9800",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)

        # Log button
        self.log_button = tk.Button(
            self.button_frame,
            text="Save To Log",
            command=self.save_log,
            width=15,
            height=2,
            bg="#9C27B0",
            fg="white",
            font=("Arial", 10, "bold"),
            disabledforeground="grey"
        )
        self.log_button.pack(side=tk.LEFT, padx=5)
        self.log_button.config(state=tk.DISABLED)

        # Open Log button
        self.open_log_button = tk.Button(
            self.button_frame,
            text="Open Log",
            command=self.open_log,
            width=15,
            height=2,
            bg="#607D8B",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.open_log_button.pack(side=tk.RIGHT, padx=5)

        # Graph button
        self.graph_button = tk.Button(
            self.button_frame,
            text="View Graph",
            command=self.show_speed_graph,
            width=15,
            height=2,
            bg="#00ACC1",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.graph_button.pack(side=tk.RIGHT, padx=5)

        # Error Log button
        self.error_log_button = tk.Button(
            self.button_frame,
            text="Error Log",
            command=self.show_error_log,
            width=15,
            height=2,
            bg="#E53935",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        self.error_log_button.pack(side=tk.RIGHT, padx=5)

        # Exit button
        self.exit_button = tk.Button(
            self.button_frame,
            text="Exit",
            command=self.root.quit,
            width=15,
            height=2,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.exit_button.pack(side=tk.LEFT, padx=5)

        self.test_running = False
        self.stop_requested = False
        self.latest_results = None
        self.tls_warning_shown = False
        self.diagnostic_queue = queue.Queue()
        self.diagnostic_window = None
        self.diagnostic_text = None
        self.diagnostic_running = False
        self.ookla_process = None
        self.apply_dark_mode()
        self.root.after(100, self.drain_diagnostic_queue)

    def apply_dark_mode(self):
        """Apply a dark color theme to the Tkinter controls."""
        theme = self.DARK_THEME
        self.root.configure(bg=theme["window"])

        for frame in (self.output_frame, self.comment_frame, self.button_frame):
            frame.configure(bg=theme["window"])

        self.comment_label.configure(
            bg=theme["window"],
            fg=theme["muted"],
            activebackground=theme["window"],
            activeforeground=theme["text"],
        )
        self.comment_entry.configure(
            bg=theme["field"],
            fg=theme["text"],
            insertbackground=theme["text"],
            selectbackground=theme["select_bg"],
            selectforeground=theme["select_fg"],
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=theme["border"],
            highlightcolor=theme["select_bg"],
        )
        self.output_text.configure(
            bg=theme["field"],
            fg=theme["text"],
            insertbackground=theme["text"],
            selectbackground=theme["select_bg"],
            selectforeground=theme["select_fg"],
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=theme["border"],
            highlightcolor=theme["select_bg"],
        )

        for button in (
            self.start_button,
            self.retest_button,
            self.multi_test_button,
            self.stop_button,
            self.log_button,
            self.open_log_button,
            self.graph_button,
            self.error_log_button,
            self.exit_button,
        ):
            button.configure(
                fg="white",
                activeforeground="white",
                disabledforeground=theme["disabled"],
                relief=tk.FLAT,
                borderwidth=0,
                highlightthickness=0,
                cursor="hand2",
            )

        self.start_button.configure(activebackground="#3f9a44")
        self.retest_button.configure(activebackground="#1976d2")
        self.multi_test_button.configure(activebackground="#283593")
        self.stop_button.configure(activebackground="#ef8f00")
        self.log_button.configure(activebackground="#8e24aa")
        self.open_log_button.configure(activebackground="#546e7a")
        self.graph_button.configure(activebackground="#0097a7")
        self.error_log_button.configure(activebackground="#b71c1c")
        self.exit_button.configure(activebackground="#d83a30")

    def extract_error_code(self, error_string):
        """Extract HTTP error code from error message"""
        import re
        match = re.search(r'(\d{3})', error_string)
        if match:
            return match.group(1)
        return 'UNKNOWN'

    def log_output(self, message):
        """Add message to output text area"""
        self.output_text.insert(tk.END, message + "\n")
        self.output_text.see(tk.END)
        self.root.update()

    def log_verbose_error(self, error):
        """Print detailed error diagnostics to the GUI output."""
        self.log_output("\n" + "=" * 60)
        self.log_output("VERBOSE ERROR DETAILS")
        self.log_output("=" * 60)
        self.log_output(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_output(f"Error Type: {type(error).__name__}")
        self.log_output(f"Error Message: {error}")

        cause = error.__cause__ or error.__context__
        if cause:
            self.log_output(f"Underlying Type: {type(cause).__name__}")
            self.log_output(f"Underlying Message: {cause}")

        self.log_output("")
        self.log_output("Traceback:")
        for line in traceback.format_exception(type(error), error, error.__traceback__):
            for output_line in line.rstrip().splitlines():
                self.log_output(output_line)
        self.log_output("=" * 60)

    def show_error_log(self):
        """Run diagnostics and show only that run in a separate window."""
        if self.test_running:
            self.log_output("[!] Cannot run diagnostics while a test is running.")
            return
        if self.diagnostic_running:
            self.log_output("[!] A diagnostic run is already in progress.")
            return
        self.diagnostic_running = True
        thread = threading.Thread(
            target=self.run_diagnostics,
            kwargs={"reason": "manual"},
        )
        thread.daemon = True
        thread.start()

    def open_diagnostic_window(self, title):
        """Create or reset the window used for one diagnostic run."""
        if self.diagnostic_window is None or not self.diagnostic_window.winfo_exists():
            self.diagnostic_window = tk.Toplevel(self.root)
            self.diagnostic_window.geometry("900x620")
            self.diagnostic_window.configure(bg=self.DARK_THEME["window"])
            self.diagnostic_text = scrolledtext.ScrolledText(
                self.diagnostic_window, wrap=tk.WORD, font=("Courier", 10),
                bg=self.DARK_THEME["field"], fg=self.DARK_THEME["text"],
                insertbackground=self.DARK_THEME["text"], relief=tk.FLAT,
                borderwidth=0,
            )
            self.diagnostic_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.diagnostic_window.title(title)
        self.diagnostic_text.delete("1.0", tk.END)
        self.diagnostic_window.deiconify()
        self.diagnostic_window.lift()

    def drain_diagnostic_queue(self):
        """Apply worker diagnostic events to Tk widgets on the main thread."""
        try:
            while True:
                action, value = self.diagnostic_queue.get_nowait()
                if action == "start":
                    self.open_diagnostic_window(value)
                elif self.diagnostic_text is not None:
                    self.diagnostic_text.insert(tk.END, value + "\n")
                    self.diagnostic_text.see(tk.END)
        except queue.Empty:
            pass
        self.root.after(100, self.drain_diagnostic_queue)

    def diagnostic_output(self, message):
        """Send one result to the diagnostic window and current-run log."""
        self.diagnostic_queue.put(("message", message))
        filename = os.path.join(os.path.dirname(__file__), 'internet_speed_diagnostics.log')
        try:
            with open(filename, 'a', encoding='utf-8') as log_file:
                log_file.write(message + "\n")
        except Exception as e:
            if not getattr(self, '_diagnostic_file_failed', False):
                self._diagnostic_file_failed = True
                self.diagnostic_queue.put(
                    ("message", f"[!] Could not write diagnostic log: {e}")
                )

    def diagnostic_header(self, title):
        """Print a consistent report section header."""
        self.diagnostic_output("")
        self.diagnostic_output("=" * 60)
        self.diagnostic_output(title)
        self.diagnostic_output("=" * 60)

    def start_diagnostic_log(self, reason):
        """Reset the display and file so they contain only the current run."""
        self._diagnostic_file_failed = False
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title = f"Error Log - {timestamp} - {reason}"
        self.diagnostic_queue.put(("start", title))
        filename = os.path.join(os.path.dirname(__file__), 'internet_speed_diagnostics.log')
        try:
            with open(filename, 'w', encoding='utf-8') as log_file:
                log_file.write("#" * 60 + "\n")
                log_file.write(f"DIAGNOSTIC RUN - {timestamp} - reason={reason}\n")
                log_file.write("#" * 60 + "\n")
        except Exception as e:
            self._diagnostic_file_failed = True
            self.diagnostic_queue.put(
                ("message", f"[!] Could not initialize diagnostic log: {e}")
            )

    def diagnose_context(self, reason, download_speed, upload_speed, ping, jitter):
        """Log the reason and any available speed-test measurements."""
        self.diagnostic_header("A) CONTEXT")
        self.diagnostic_output(f"Reason: {reason}")
        self.diagnostic_output(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        measurements = (("Download", download_speed, "Mbps"),
                        ("Upload", upload_speed, "Mbps"),
                        ("Ping", ping, "ms"), ("Jitter", jitter, "ms"))
        for label, value, unit in measurements:
            if value is not None:
                self.diagnostic_output(f"{label}: {value:.2f} {unit}")
        if download_speed is not None:
            delta = download_speed - self.EXPECTED_MIN_DOWNLOAD_MBPS
            self.diagnostic_output(f"Download delta vs expected: {delta:+.2f} Mbps")
        if upload_speed is not None:
            delta = upload_speed - self.EXPECTED_MIN_UPLOAD_MBPS
            self.diagnostic_output(f"Upload delta vs expected: {delta:+.2f} Mbps")

    def diagnose_system_stack(self):
        """Log platform, hostname, interfaces, and selected default route."""
        self.diagnostic_header("B) SYSTEM & NETWORK STACK")
        try:
            self.diagnostic_output(f"Platform: {platform.platform()}")
            self.diagnostic_output(f"Python: {platform.python_version()}")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Platform details: {e}")
        try:
            hostname = socket.gethostname()
            self.diagnostic_output(f"Hostname: {hostname}")
            self.diagnostic_output(f"Primary hostname address: {socket.gethostbyname(hostname)}")
            addresses = sorted({item[4][0] for item in socket.getaddrinfo(hostname, None)})
            self.diagnostic_output(f"Local interface addresses: {', '.join(addresses)}")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Local address discovery: {e}")
        route_socket = None
        try:
            route_socket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            route_socket.settimeout(5)
            route_socket.connect(("1.1.1.1", 80))
            self.diagnostic_output(f"Active default-route interface: {route_socket.getsockname()[0]}")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Default route check: {e}")
        finally:
            if route_socket:
                route_socket.close()

    def diagnose_dns(self, flags):
        """Time public DNS lookups and identify suspicious resolutions."""
        self.diagnostic_header("C) DNS RESOLUTION TIMING")
        for host, port in self.DIAGNOSTIC_HOSTS:
            try:
                started = time.perf_counter()
                info = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
                elapsed_ms = (time.perf_counter() - started) * 1000
                addresses = sorted({item[4][0] for item in info})
                status = " SLOW" if elapsed_ms > 200 else ""
                self.diagnostic_output(
                    f"{host}: {elapsed_ms:.1f} ms{status}; A records={', '.join(addresses)}"
                )
                if elapsed_ms > 200:
                    flags['slow_dns'].append((host, elapsed_ms))
                suspicious = [address for address in addresses
                              if ipaddress.ip_address(address).is_private
                              or ipaddress.ip_address(address).is_loopback]
                if suspicious:
                    flags['dns_redirects'].append((host, suspicious))
                    self.diagnostic_output(
                        f"[!] Likely DNS hijack/proxy redirect for {host}: {', '.join(suspicious)}"
                    )
            except Exception as e:
                self.diagnostic_output(f"[FAIL] DNS {host}: {e}")

    def diagnose_tcp(self, flags):
        """Time raw TCP connections to every diagnostic target."""
        self.diagnostic_header("D) TCP CONNECT TIMING")
        for host, port in self.DIAGNOSTIC_HOSTS:
            tcp_socket = None
            try:
                tcp_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                tcp_socket.settimeout(5)
                started = time.perf_counter()
                tcp_socket.connect((host, port))
                elapsed_ms = (time.perf_counter() - started) * 1000
                status = " SLOW" if elapsed_ms > 300 else " OK"
                self.diagnostic_output(f"{host}:{port}: {elapsed_ms:.1f} ms{status}")
                flags['tcp_ok'].add(host)
            except Exception as e:
                flags['tcp_fail'].add(host)
                self.diagnostic_output(f"[FAIL] TCP {host}:{port}: {e}")
            finally:
                if tcp_socket:
                    tcp_socket.close()

    @staticmethod
    def certificate_common_name(name_parts):
        """Extract a certificate commonName from ssl.getpeercert output."""
        for group in name_parts:
            for key, value in group:
                if key == 'commonName':
                    return value
        return "Unknown"

    def diagnose_tls(self, flags):
        """Inspect TLS negotiation and detect well-known interception issuers."""
        self.diagnostic_header("E) TLS HANDSHAKE INSPECTION")
        inspectors = ("Norton", "Symantec", "NortonLifeLock", "Kaspersky", "ESET",
                      "Bitdefender", "Fortinet", "Zscaler", "Bluecoat")
        for host, port in self.DIAGNOSTIC_HOSTS:
            try:
                context = ssl.create_default_context()
                with socket.create_connection((host, port), timeout=5) as raw_socket:
                    raw_socket.settimeout(5)
                    started = time.perf_counter()
                    with context.wrap_socket(raw_socket, server_hostname=host) as tls_socket:
                        elapsed_ms = (time.perf_counter() - started) * 1000
                        cert = tls_socket.getpeercert()
                        subject = self.certificate_common_name(cert.get('subject', ()))
                        issuer = self.certificate_common_name(cert.get('issuer', ()))
                        self.diagnostic_output(f"{host}: handshake={elapsed_ms:.1f} ms")
                        self.diagnostic_output(f"    Subject CN: {subject}")
                        self.diagnostic_output(f"    Issuer CN: {issuer}")
                        self.diagnostic_output(
                            f"    Valid: {cert.get('notBefore', 'Unknown')} to {cert.get('notAfter', 'Unknown')}"
                        )
                        self.diagnostic_output(
                            f"    TLS: {tls_socket.version()}; Cipher: {tls_socket.cipher()}"
                        )
                        detected = next((name for name in inspectors
                                         if name.lower() in issuer.lower()), None)
                        if detected:
                            flags['tls_interceptors'].add(issuer)
                            self.diagnostic_output(
                                f"[!] TLS chain is being intercepted by {issuer}. This breaks"
                            )
                            self.diagnostic_output(
                                "    Cloudflare's edge fingerprinting and throttles throughput."
                            )
            except Exception as e:
                self.diagnostic_output(f"[FAIL] TLS {host}:{port}: {e}")

    def diagnose_http(self, flags):
        """Probe Cloudflare metadata and log middlebox-sensitive headers."""
        self.diagnostic_header("F) HTTP DIAGNOSTIC PROBE")
        try:
            body, response = self.cloudflare_request("/meta", timeout=10)
            self.diagnostic_output(f"Cloudflare /meta: {len(body)} bytes")
            header_names = ('server', 'cf-ray', 'cf-cache-status', 'alt-svc', 'x-amz-cf-id')
            response_headers = {key.lower(): value for key, value in response.headers.items()}
            for name in header_names:
                self.diagnostic_output(f"{name}: {response_headers.get(name, '<missing>')}")
            if not response_headers.get('cf-ray'):
                flags['missing_cf_ray'] = True
                self.diagnostic_output(
                    "[!] Missing cf-ray header; an inspecting middlebox may have stripped it."
                )
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Cloudflare HTTP probe: {e}")

    def diagnose_proxy_vpn(self, flags):
        """Report proxy configuration and suspicious virtual interfaces."""
        self.diagnostic_header("G) PROXY / VPN DETECTION")
        for name in ('HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY', 'NO_PROXY'):
            self.diagnostic_output(f"{name}: {os.environ.get(name, '<not set>')}")
        try:
            self.diagnostic_output(f"urllib proxies: {urllib.request.getproxies() or '<none>'}")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Proxy discovery: {e}")
        products = ("Norton", "Nord", "Tailscale", "WireGuard", "TAP", "OpenVPN",
                    "Proton", "Mullvad", "Cloudflare WARP", "NDivert", "Zscaler")
        try:
            import psutil
            interfaces = psutil.net_if_addrs()
            if not interfaces:
                self.diagnostic_output("Network interfaces: <none reported>")
            for name, addresses in interfaces.items():
                values = [item.address for item in addresses if item.address]
                self.diagnostic_output(f"Interface {name}: {', '.join(values)}")
                if any(product.lower() in name.lower() for product in products):
                    flags['vpn_interfaces'].add(name)
                    self.diagnostic_output(f"[!] VPN/security interface detected: {name}")
        except ImportError:
            self.diagnostic_output("psutil not installed; detailed interface names unavailable.")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Network interface discovery: {e}")

    def diagnose_recommendations(self, flags, download_speed, upload_speed):
        """Turn detected conditions into plain-English next steps."""
        self.diagnostic_header("H) RECOMMENDATIONS")
        recommendations = []
        if flags['tls_interceptors']:
            issuers = ', '.join(sorted(flags['tls_interceptors']))
            recommendations.append(
                f"TLS inspection is rewriting certificates ({issuers}). Disable HTTPS Scanning "
                "in the security product, then re-run the test."
            )
        if flags['vpn_interfaces']:
            names = ', '.join(sorted(flags['vpn_interfaces']))
            recommendations.append(
                f"VPN/security interfaces were detected ({names}). Disconnect the VPN and retry."
            )
        if any(elapsed > 500 for _, elapsed in flags['slow_dns']):
            recommendations.append(
                "DNS resolution exceeded 500 ms. Try DNS servers 1.1.1.1 or 8.8.8.8."
            )
        if '1.1.1.1' in flags['tcp_fail'] and 'www.google.com' in flags['tcp_ok']:
            recommendations.append(
                "Cloudflare failed while Google connected. Check ISP or corporate-firewall blocking."
            )
        if flags['dns_redirects']:
            recommendations.append(
                "A public host resolved privately. Check DNS filtering, proxy, and router settings."
            )
        if flags['missing_cf_ray']:
            recommendations.append(
                "Cloudflare's cf-ray header was missing. Temporarily disable web inspection and retry."
            )
        if ((download_speed is not None and download_speed < self.EXPECTED_MIN_DOWNLOAD_MBPS)
                or (upload_speed is not None and upload_speed < self.EXPECTED_MIN_UPLOAD_MBPS)):
            recommendations.append(
                "Measured speed is below the configured threshold. Test by Ethernet and pause other traffic."
            )
        if not recommendations:
            recommendations.append(
                "No obvious DNS, TCP, TLS, proxy, or VPN issue was detected. Retry at another time."
            )
        for number, recommendation in enumerate(recommendations, 1):
            self.diagnostic_output(f"{number}. {recommendation}")

    def run_diagnostics(self, reason, download_speed=None, upload_speed=None,
                        ping=None, jitter=None):
        """Collect a resilient, persistent network-underperformance report."""
        self.diagnostic_running = True
        try:
            self.start_diagnostic_log(reason)
            flags = {
                'tls_interceptors': set(),
                'vpn_interfaces': set(),
                'slow_dns': [],
                'tcp_ok': set(),
                'tcp_fail': set(),
                'dns_redirects': [],
                'missing_cf_ray': False,
            }
            self.diagnose_context(reason, download_speed, upload_speed, ping, jitter)
            self.diagnose_system_stack()
            self.diagnose_dns(flags)
            self.diagnose_tcp(flags)
            self.diagnose_tls(flags)
            self.diagnose_http(flags)
            self.diagnose_proxy_vpn(flags)
            self.diagnose_recommendations(flags, download_speed, upload_speed)
            self.diagnostic_output("Diagnostic run complete.")
        except Exception as e:
            self.diagnostic_output(f"[FAIL] Diagnostic run failed unexpectedly: {e}")
            for line in traceback.format_exception(type(e), e, e.__traceback__):
                for output_line in line.rstrip().splitlines():
                    self.diagnostic_output(output_line)
        finally:
            self.diagnostic_running = False

    def open_log(self):
        """Open the Excel log file in the default system app."""
        log_dir = os.path.dirname(__file__)
        filename = os.path.join(log_dir, 'internet_speed_log.xlsx')
        if not os.path.exists(filename):
            filename = os.path.join(log_dir, 'internet_speed_log.csv')

        if not os.path.exists(filename):
            self.log_output(f"Log file not found yet: {filename}")
            self.log_output("Run a test and click 'Save To Log' first.")
            return

        try:
            os.startfile(filename)
        except Exception as e:
            self.log_output(f"Failed to open log: {e}")

    def load_graph_points(self):
        """Load graphable speed-test rows from the CSV log."""
        filename = os.path.join(os.path.dirname(__file__), 'internet_speed_log.csv')
        if not os.path.exists(filename):
            raise FileNotFoundError(filename)

        points = []
        with open(filename, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    timestamp = datetime.strptime(row.get('timestamp', ''), '%Y-%m-%d %H:%M:%S')
                    points.append({
                        'timestamp': timestamp,
                        'download_mbps': float(row.get('download_mbps', 0)),
                        'upload_mbps': float(row.get('upload_mbps', 0)),
                        'ping_ms': float(row.get('ping_ms', 0)),
                        'jitter_ms': float(row.get('jitter_ms', 0)),
                    })
                except (TypeError, ValueError):
                    continue

        points.sort(key=lambda point: point['timestamp'])
        return points

    def show_speed_graph(self):
        """Open a dark-mode line graph of the speed log."""
        try:
            points = self.load_graph_points()
        except FileNotFoundError as e:
            self.log_output(f"Graph log file not found: {e.filename}")
            messagebox.showinfo("No Speed Log", "Run tests and save them to the log before viewing the graph.")
            return
        except Exception as e:
            self.log_output(f"Failed to read graph data: {e}")
            messagebox.showerror("Graph Error", f"Could not read the speed log:\n{e}")
            return

        if not points:
            self.log_output("No valid speed log rows found for graphing.")
            messagebox.showinfo("No Graph Data", "No valid speed log rows were found.")
            return

        graph_window = tk.Toplevel(self.root)
        graph_window.title("Speed Log Graph v1.08")
        graph_window.geometry("1100x760")
        graph_window.minsize(820, 620)
        graph_window.configure(bg=self.GRAPH_COLORS["background"])

        header = tk.Frame(graph_window, bg=self.GRAPH_COLORS["background"])
        header.pack(fill=tk.X, padx=18, pady=(14, 4))

        title = tk.Label(
            header,
            text="Internet Speed Log Telemetry",
            bg=self.GRAPH_COLORS["background"],
            fg=self.GRAPH_COLORS["text"],
            font=("Segoe UI", 17, "bold")
        )
        title.pack(side=tk.LEFT)

        subtitle = tk.Label(
            header,
            text=f"{len(points)} logged tests",
            bg=self.GRAPH_COLORS["background"],
            fg=self.GRAPH_COLORS["muted"],
            font=("Segoe UI", 10)
        )
        subtitle.pack(side=tk.RIGHT)

        canvas = tk.Canvas(
            graph_window,
            bg=self.GRAPH_COLORS["background"],
            highlightthickness=0,
            bd=0
        )
        canvas.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4, 16))
        canvas.bind("<Configure>", lambda event: self.draw_speed_graph(canvas, points))
        self.draw_speed_graph(canvas, points)
        self.log_output("Opened dark-mode speed log graph.")

    def draw_speed_graph(self, canvas, points):
        """Draw separate speed and latency charts on a Tkinter canvas."""
        colors = self.GRAPH_COLORS
        canvas.delete("all")

        width = max(canvas.winfo_width(), 760)
        height = max(canvas.winfo_height(), 560)

        canvas.create_rectangle(0, 0, width, height, fill=colors["background"], outline="")
        canvas.create_rectangle(
            18, 18, width - 18, height - 18,
            fill=colors["panel"],
            outline="#1b2d46",
            width=1
        )

        speed_series = [
            ("download_mbps", "Download Mbps", colors["download"]),
            ("upload_mbps", "Upload Mbps", colors["upload"]),
        ]
        latency_series = [
            ("ping_ms", "Ping ms", colors["ping"]),
            ("jitter_ms", "Jitter ms", colors["jitter"]),
        ]
        top_bounds = (82, 52, width - 34, (height - 92) * 0.52)
        bottom_bounds = (82, top_bounds[3] + 58, width - 34, height - 92)

        self.draw_graph_panel(
            canvas,
            points,
            speed_series,
            top_bounds,
            "Internet Speed",
            "Mbps",
            show_x_labels=False
        )
        self.draw_graph_panel(
            canvas,
            points,
            latency_series,
            bottom_bounds,
            "Ping and Jitter",
            "ms",
            show_x_labels=True
        )

        latest = points[-1]
        latest_text = (
            f"Latest  Download {latest['download_mbps']:.2f} Mbps"
            f"   Upload {latest['upload_mbps']:.2f} Mbps"
            f"   Ping {latest['ping_ms']:.2f} ms"
            f"   Jitter {latest['jitter_ms']:.2f} ms"
        )
        canvas.create_text(
            82, height - 58,
            text=latest_text,
            fill=colors["text"],
            font=("Segoe UI", 10, "bold"),
            anchor=tk.W
        )

    def draw_graph_panel(self, canvas, points, series, bounds, title, unit, show_x_labels):
        """Draw one chart panel using all loaded speed log rows."""
        colors = self.GRAPH_COLORS
        plot_left, plot_top, plot_right, plot_bottom = bounds
        plot_width = max(plot_right - plot_left, 1)
        plot_height = max(plot_bottom - plot_top, 1)
        values = [point[key] for point in points for key, _, _ in series]
        y_max = self.nice_axis_max(max(values) if values else 1)

        def x_for(index):
            if len(points) == 1:
                return plot_left + plot_width / 2
            return plot_left + (index / (len(points) - 1)) * plot_width

        def y_for(value):
            return plot_bottom - (value / y_max) * plot_height

        canvas.create_text(
            plot_left, plot_top - 24,
            text=f"{title} ({unit})",
            fill=colors["text"],
            font=("Segoe UI", 12, "bold"),
            anchor=tk.W
        )

        legend_width = 150
        legend_x = plot_right - (legend_width * len(series))
        legend_y = plot_top - 24
        for offset, (_, label, color) in enumerate(series):
            x = legend_x + offset * legend_width
            canvas.create_line(x, legend_y, x + 28, legend_y, fill=color, width=4)
            canvas.create_text(
                x + 36, legend_y,
                text=label,
                fill=colors["text"],
                font=("Segoe UI", 9, "bold"),
                anchor=tk.W
            )

        for i in range(6):
            value = y_max * i / 5
            y = plot_bottom - plot_height * i / 5
            line_color = colors["axis"] if i == 0 else colors["grid"]
            canvas.create_line(plot_left, y, plot_right, y, fill=line_color, width=1)
            canvas.create_text(
                plot_left - 14, y,
                text=f"{value:g}",
                fill=colors["muted"],
                font=("Segoe UI", 9),
                anchor=tk.E
            )

        tick_indexes = self.graph_tick_indexes(len(points), 7)
        for index in tick_indexes:
            x = x_for(index)
            timestamp = points[index]["timestamp"]
            canvas.create_line(x, plot_top, x, plot_bottom, fill=colors["grid"], width=1)
            if show_x_labels:
                canvas.create_text(
                    x, plot_bottom + 26,
                    text=timestamp.strftime("%m-%d\n%H:%M"),
                    fill=colors["muted"],
                    font=("Segoe UI", 9),
                    anchor=tk.N
                )

        canvas.create_line(plot_left, plot_top, plot_left, plot_bottom, fill=colors["axis"], width=2)
        canvas.create_line(plot_left, plot_bottom, plot_right, plot_bottom, fill=colors["axis"], width=2)

        for key, _, color in series:
            coordinates = []
            for index, point in enumerate(points):
                coordinates.extend([x_for(index), y_for(point[key])])

            if len(points) == 1:
                x = coordinates[0]
                y = coordinates[1]
                canvas.create_oval(x - 5, y - 5, x + 5, y + 5, outline=color, fill=color)
                canvas.create_text(x + 10, y - 8, text=f"{points[0][key]:.2f}", fill=color, anchor=tk.W)
                continue

            canvas.create_line(*coordinates, fill="#17283c", width=9, smooth=True)
            canvas.create_line(*coordinates, fill=color, width=3, smooth=True)

            for index in self.graph_marker_indexes(len(points)):
                x = x_for(index)
                y = y_for(points[index][key])
                canvas.create_oval(x - 3, y - 3, x + 3, y + 3, outline=color, fill=colors["panel"], width=2)

        if show_x_labels:
            canvas.create_text(
                (plot_left + plot_right) / 2, plot_bottom + 62,
                text="Timestamp",
                fill=colors["text"],
                font=("Segoe UI", 10, "bold")
            )

    def graph_tick_indexes(self, point_count, max_ticks):
        """Return evenly spaced indexes while still using all data for the line."""
        if point_count <= 0:
            return []
        if point_count == 1:
            return [0]
        tick_count = min(max_ticks, point_count)
        return sorted(set(
            round(i * (point_count - 1) / max(tick_count - 1, 1))
            for i in range(tick_count)
        ))

    def graph_marker_indexes(self, point_count):
        """Keep markers readable without dropping any line data."""
        if point_count <= 18:
            return list(range(point_count))
        step = max(point_count // 8, 1)
        indexes = list(range(0, point_count, step))
        if indexes[-1] != point_count - 1:
            indexes.append(point_count - 1)
        return indexes

    def nice_axis_max(self, value):
        """Round the graph Y-axis ceiling up to a readable value."""
        if value <= 0:
            return 1

        magnitude = 1
        while magnitude * 10 < value:
            magnitude *= 10

        for multiplier in (1, 2, 5, 10):
            candidate = multiplier * magnitude
            if candidate >= value:
                return candidate

        return magnitude * 10

    def write_excel_log(self, filename, header, rows):
        """Write log rows as a formatted Excel table."""
        if Workbook is None:
            raise ImportError("openpyxl is required to write formatted Excel logs")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Speed Test Log"

        worksheet.append(header)
        for row in rows:
            worksheet.append(row)

        table_end_column = get_column_letter(len(header))
        table_end_row = max(len(rows) + 1, 2)
        table_ref = f"A1:{table_end_column}{table_end_row}"
        table = Table(displayName="SpeedTestLog", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium13",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

        for column_index, column_name in enumerate(header, start=1):
            values = [column_name]
            values.extend(str(row[column_index - 1]) for row in rows if len(row) >= column_index)
            width = min(max(len(value) for value in values) + 2, 45)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        workbook.save(filename)

    def stop_test(self):
        """Stop the running test"""
        if not self.test_running:
            self.output_text.delete(1.0, tk.END)
            self.log_output("No tests are running")
            self.log_output("")
            self.log_output("Click 'Start Test' to begin speed testing...\n")
            return

        self.stop_requested = True
        if self.ookla_process is not None and self.ookla_process.poll() is None:
            try:
                self.ookla_process.terminate()
            except Exception:
                pass
        self.log_output("\n[INFO] User stopped the tests")
        self.log_output("")
        self.log_output("Click 'Start Test' to begin speed testing...\n")

    def draw_diagonal_lines(self):
        """Draw diagonal lines on canvas for disabled state"""
        self.log_button_canvas.delete("all")
        width = 100
        height = 48
        spacing = width // 6
        for i in range(8):
            x_start = i * spacing - height
            self.log_button_canvas.create_line(
                x_start, 0,
                x_start + height, height,
                fill="grey",
                width=2
            )

    def start_test(self):
        """Start the Cloudflare speed test in a separate thread"""
        if self.test_running:
            self.log_output("Test already in progress...")
            return

        self.test_running = True
        self.start_button.config(state=tk.DISABLED)
        self.retest_button.config(state=tk.DISABLED)
        self.multi_test_button.config(state=tk.DISABLED)
        self.log_button.config(state=tk.DISABLED)

        self.output_text.delete(1.0, tk.END)
        self.log_output("=" * 80)
        self.log_output(f"Speed Test Started - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_output("=" * 80)
        self.log_output("")

        test_thread = threading.Thread(target=self.run_speedtest)
        test_thread.daemon = True
        test_thread.start()

    def build_cloudflare_request(self, path, params=None, data=None, method=None):
        """Build a Cloudflare Speed Test request with browser-like headers."""
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36 SpeedTestGUI/1.08"
            ),
            "Accept": "application/json,text/plain,*/*",
            "Referer": f"{self.CLOUDFLARE_BASE_URL}/",
            "Origin": self.CLOUDFLARE_BASE_URL,
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }
        if data is not None:
            headers["Content-Type"] = "application/octet-stream"

        query = ""
        if params:
            query = "?" + urllib.parse.urlencode(params)

        url = f"{self.CLOUDFLARE_BASE_URL}{path}{query}"
        return urllib.request.Request(url, data=data, headers=headers, method=method)

    def open_cloudflare_request(self, request, timeout, context=None):
        """Open a Cloudflare request and normalize HTTP errors."""
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
                return response.read(), response
        except urllib.error.HTTPError as e:
            raise Exception(f"HTTP {e.code} from Cloudflare: {e.reason}") from e
        except (TimeoutError, socket.timeout) as e:
            raise Exception(f"Network timeout calling Cloudflare: {e}") from e

    def cloudflare_request(self, path, params=None, data=None, method=None, timeout=30):
        """Make one request to Cloudflare Speed Test."""
        if self.stop_requested:
            raise Exception("Test stopped by user")

        request_method = method or ("POST" if data is not None else "GET")
        payload_size = len(data) if data is not None else 0
        request_details = (
            f"{request_method} {path}, timeout={timeout}s, "
            f"payload={payload_size / 1_000_000:.2f} MB"
        )
        request = self.build_cloudflare_request(path, params, data, method)

        try:
            return self.open_cloudflare_request(request, timeout)
        except urllib.error.URLError as e:
            if isinstance(e.reason, ssl.SSLCertVerificationError):
                if not self.tls_warning_shown:
                    self.log_output("[WARN] TLS certificate verification failed; retrying without certificate verification.")
                    self.tls_warning_shown = True
                context = ssl._create_unverified_context()
                request = self.build_cloudflare_request(path, params, data, method)
                return self.open_cloudflare_request(request, timeout, context)
            raise Exception(f"Network error calling Cloudflare ({request_details}): {e.reason}") from e
        except (TimeoutError, socket.timeout) as e:
            raise Exception(f"Network timeout calling Cloudflare ({request_details}): {e}") from e

    def get_cloudflare_meta(self):
        """Fetch Cloudflare metadata for client and edge details."""
        try:
            body, _ = self.cloudflare_request("/meta", timeout=15)
            return json.loads(body.decode("utf-8"))
        except Exception as e:
            self.log_output(f"[WARN] Cloudflare /meta failed ({e}); using /cdn-cgi/trace fallback.")
            body, _ = self.cloudflare_request("/cdn-cgi/trace", timeout=15)
            trace = {}
            for line in body.decode("utf-8", errors="replace").splitlines():
                if "=" in line:
                    key, value = line.split("=", 1)
                    trace[key] = value
            return {
                "hostname": trace.get("h", "speed.cloudflare.com"),
                "clientIp": trace.get("ip", "N/A"),
                "httpProtocol": trace.get("http", "N/A"),
                "asOrganization": "N/A",
                "country": trace.get("loc", "N/A"),
                "city": "N/A",
                "region": "N/A",
                "latitude": "N/A",
                "longitude": "N/A",
                "colo": trace.get("colo", "Cloudflare"),
            }

    def measure_latency(self, samples=6):
        """Measure latency and jitter with lightweight Cloudflare requests."""
        latencies = []
        self.log_output("Calculating ping and jitter...")

        for i in range(samples):
            if self.stop_requested:
                raise Exception("Test stopped by user")
            start = time.perf_counter()
            self.cloudflare_request(
                "/__down",
                params={"bytes": 0, "cachebust": f"{time.time_ns()}-{i}"},
                timeout=10
            )
            elapsed_ms = (time.perf_counter() - start) * 1000
            latencies.append(elapsed_ms)

        ping = min(latencies)
        jitter = max(latencies) - min(latencies) if len(latencies) > 1 else 0.0
        return ping, jitter

    def measure_download(self):
        """Measure download speed through Cloudflare __down."""
        total_bytes = 0
        total_seconds = 0.0
        sizes = [1_000_000, 5_000_000, 10_000_000, 25_000_000]

        for size in sizes:
            if self.stop_requested:
                raise Exception("Test stopped by user")
            self.log_output(f"Downloading {size / 1_000_000:.0f} MB sample...")
            start = time.perf_counter()
            body, _ = self.cloudflare_request(
                "/__down",
                params={"bytes": size, "cachebust": time.time_ns()},
                timeout=45
            )
            elapsed = time.perf_counter() - start
            total_bytes += len(body)
            total_seconds += elapsed

        return (total_bytes * 8) / total_seconds / 1_000_000

    def measure_upload(self):
        """Measure upload speed through Cloudflare __up."""
        total_bytes = 0
        total_seconds = 0.0

        for size in self.UPLOAD_SAMPLE_SIZES:
            if self.stop_requested:
                raise Exception("Test stopped by user")
            sample_mb = size / 1_000_000
            payload = b"0" * size

            for attempt in range(1, self.UPLOAD_RETRIES + 2):
                try:
                    self.log_output(f"Uploading {sample_mb:.1f} MB sample...")
                    start = time.perf_counter()
                    self.cloudflare_request("/__up", data=payload, method="POST", timeout=60)
                    elapsed = time.perf_counter() - start
                    total_bytes += size
                    total_seconds += elapsed
                    break
                except Exception as e:
                    if self.stop_requested or "Test stopped by user" in str(e):
                        raise
                    if attempt > self.UPLOAD_RETRIES:
                        if total_bytes:
                            self.log_output(f"[WARN] Skipping {sample_mb:.1f} MB upload sample after timeout/error: {e}")
                            break
                        raise Exception(
                            "Upload test could not reach Cloudflare. "
                            f"Last error: {e}"
                        ) from e
                    self.log_output(
                        f"[WARN] Upload sample failed ({e}); retrying "
                        f"{attempt}/{self.UPLOAD_RETRIES}..."
                    )

        if total_seconds <= 0:
            raise Exception("Upload test did not complete any samples.")

        return (total_bytes * 8) / total_seconds / 1_000_000

    def run_speedtest(self):
        """Run the Cloudflare speed test."""
        try:
            self.log_output("Initializing Cloudflare Speed Test...")
            self.log_output("Getting Cloudflare metadata...")
            meta = self.get_cloudflare_meta()

            location_parts = [
                meta.get("city"),
                meta.get("region"),
                meta.get("country"),
            ]
            location = ", ".join(part for part in location_parts if part) or "N/A"
            lat = meta.get("latitude", "N/A")
            lon = meta.get("longitude", "N/A")
            if meta.get("loc") and "," in meta["loc"]:
                lat, lon = [part.strip() for part in meta["loc"].split(",", 1)]

            colo = meta.get("colo", "Cloudflare")
            if isinstance(colo, dict):
                server_name = colo.get("iata", "Cloudflare")
            else:
                server_name = colo
            sponsor = "Cloudflare"
            isp = meta.get("asOrganization", "N/A")
            ip_address = meta.get("clientIp", "N/A")
            country = meta.get("country", "N/A")

            self.log_output("\n" + "=" * 60)
            self.log_output("CLOUDFLARE EDGE INFORMATION")
            self.log_output("=" * 60)
            self.log_output(f"Server: {server_name}")
            self.log_output(f"Sponsor: {sponsor}")
            self.log_output(f"Location: {location}")
            self.log_output(f"Country Code: {country}")
            self.log_output(f"Latitude: {lat}")
            self.log_output(f"Longitude: {lon}")

            self.log_output("\n" + "=" * 60)
            self.log_output("ISP INFORMATION")
            self.log_output("=" * 60)
            self.log_output(f"ISP: {isp}")
            self.log_output(f"IP Address: {ip_address}")
            self.log_output(f"Country: {country}")

            self.log_output("\n" + "=" * 60)
            self.log_output("PING & LATENCY")
            self.log_output("=" * 60)
            ping, jitter = self.measure_latency()
            self.log_output(f"Ping (Latency): {ping:.2f} ms")
            self.log_output(f"Jitter: {jitter:.2f} ms")

            self.log_output("\n" + "=" * 60)
            self.log_output("DOWNLOAD SPEED TEST")
            self.log_output("=" * 60)
            download_speed = self.measure_download()
            self.log_output(f"Download Speed: {download_speed:.2f} Mbps")

            self.log_output("\n" + "=" * 60)
            self.log_output("UPLOAD SPEED TEST")
            self.log_output("=" * 60)
            upload_speed = self.measure_upload()
            self.log_output(f"Upload Speed: {upload_speed:.2f} Mbps")

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.latest_results = {
                'timestamp': timestamp,
                'server': server_name,
                'sponsor': sponsor,
                'location': location,
                'country': country,
                'cc': country,
                'lat': lat,
                'lon': lon,
                'isp': isp,
                'ip': ip_address,
                'download_mbps': round(download_speed, 2),
                'upload_mbps': round(upload_speed, 2),
                'ping_ms': round(ping, 2),
                'jitter_ms': round(jitter, 2),
                'comments': ''
            }

            self.log_output("\n" + "=" * 60)
            self.log_output("TEST SUMMARY")
            self.log_output("=" * 60)
            self.log_output(f"Download: {download_speed:.2f} Mbps")
            self.log_output(f"Upload: {upload_speed:.2f} Mbps")
            self.log_output(f"Ping: {ping:.2f} ms")
            self.log_output(f"Jitter: {jitter:.2f} ms")
            self.log_output("=" * 60)
            self.log_output(f"Test Completed - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self.log_output("=" * 60)

            if (download_speed < self.EXPECTED_MIN_DOWNLOAD_MBPS
                    or upload_speed < self.EXPECTED_MIN_UPLOAD_MBPS):
                self.log_output("")
                self.log_output("[!] Speed is below the expected threshold "
                                f"({self.EXPECTED_MIN_DOWNLOAD_MBPS} Mbps down / "
                                f"{self.EXPECTED_MIN_UPLOAD_MBPS} Mbps up).")
                self.log_output("    Running automatic diagnostics...")
                self.run_diagnostics(reason="slow_speed",
                                     download_speed=download_speed,
                                     upload_speed=upload_speed,
                                     ping=ping,
                                     jitter=jitter)

            self.log_button.config(state=tk.NORMAL)

        except Exception as e:
            self.log_output(f"\nERROR: {str(e)}")
            if "Test stopped by user" not in str(e):
                self.log_verbose_error(e)
                self.log_output("Please check your internet connection and try again.")

        finally:
            self.test_running = False
            self.stop_requested = False
            self.start_button.config(state=tk.NORMAL)
            self.retest_button.config(state=tk.NORMAL)
            self.multi_test_button.config(state=tk.NORMAL)

    # =========================================================================
    # MULTI-PROVIDER TESTING (v1.08)
    # =========================================================================

    def start_multi_test(self):
        """Kick off the multi-provider test in a background thread."""
        if self.test_running:
            self.log_output("Test already in progress...")
            return

        self.test_running = True
        self.start_button.config(state=tk.DISABLED)
        self.retest_button.config(state=tk.DISABLED)
        self.multi_test_button.config(state=tk.DISABLED)
        self.log_button.config(state=tk.DISABLED)

        self.output_text.delete(1.0, tk.END)
        self.log_output("=" * 80)
        self.log_output(f"Multi-Provider Test Started - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_output("=" * 80)
        self.log_output("Running Cloudflare, Fast.com (Netflix), and Ookla (if installed)...")
        self.log_output("This will take 2-3 minutes. Press 'Stop' to abort.\n")

        thread = threading.Thread(target=self.run_multi_provider_test, daemon=True)
        thread.start()

    def run_multi_provider_test(self):
        """Run every available provider sequentially and compare results."""
        results = []
        try:
            self.log_output("\n[1/3] Cloudflare speed test...")
            self.log_output("-" * 60)
            try:
                cf_result = self.measure_cloudflare()
                results.append(cf_result)
            except Exception as e:
                if "Test stopped by user" in str(e):
                    raise
                self.log_output(f"[FAIL] Cloudflare test failed: {e}")
                results.append({"provider": "Cloudflare", "error": str(e)})

            if self.stop_requested:
                raise Exception("Test stopped by user")

            self.log_output("\n[2/3] Netflix Fast.com test...")
            self.log_output("-" * 60)
            try:
                fast_result = self.measure_fastdotcom()
                results.append(fast_result)
            except Exception as e:
                if "Test stopped by user" in str(e):
                    raise
                self.log_output(f"[FAIL] Fast.com test failed: {e}")
                results.append({"provider": "Fast.com", "error": str(e)})

            if self.stop_requested:
                raise Exception("Test stopped by user")

            self.log_output("\n[3/3] Ookla speedtest CLI...")
            self.log_output("-" * 60)
            ookla_result = self.measure_ookla_cli()
            if ookla_result is None:
                self.log_output("[SKIP] Ookla speedtest CLI is not installed.")
                self.log_output("       Install from: https://www.speedtest.net/apps/cli")
                self.log_output("       Then re-run Multi-Test to add Ookla to the comparison.")
            else:
                results.append(ookla_result)

            self.show_multi_provider_comparison(results)
            self.save_multi_provider_log(results)

        except Exception as e:
            if "Test stopped by user" in str(e):
                self.log_output("\n[INFO] Multi-test stopped by user.")
            else:
                self.log_output(f"\nERROR: {e}")
                self.log_verbose_error(e)
        finally:
            self.test_running = False
            self.stop_requested = False
            self.ookla_process = None
            self.start_button.config(state=tk.NORMAL)
            self.retest_button.config(state=tk.NORMAL)
            self.multi_test_button.config(state=tk.NORMAL)

    def measure_cloudflare(self):
        """Run a Cloudflare measurement and return it as a provider result."""
        meta = self.get_cloudflare_meta()
        colo = meta.get("colo", "Cloudflare")
        if isinstance(colo, dict):
            server_name = colo.get("iata", "Cloudflare")
        else:
            server_name = colo
        location_parts = [meta.get("city"), meta.get("country")]
        location = ", ".join(p for p in location_parts if p) or "Cloudflare"
        server = f"{server_name} ({location})"

        self.log_output(f"Cloudflare edge: {server}")
        ping, jitter = self.measure_latency()
        self.log_output(f"Ping: {ping:.2f} ms, Jitter: {jitter:.2f} ms")
        download_speed = self.measure_download()
        self.log_output(f"Download: {download_speed:.2f} Mbps")
        upload_speed = self.measure_upload()
        self.log_output(f"Upload: {upload_speed:.2f} Mbps")

        return {
            "provider": "Cloudflare",
            "download_mbps": download_speed,
            "upload_mbps": upload_speed,
            "ping_ms": ping,
            "jitter_ms": jitter,
            "server": server,
            "isp": meta.get("asOrganization", "N/A"),
        }

    def fastdotcom_get_targets(self):
        """Fetch the Netflix Fast.com target list."""
        url = self.FASTDOTCOM_API.format(
            token=self.FASTDOTCOM_TOKEN,
            count=self.FAST_PARALLEL_STREAMS,
        )
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36 SpeedTestGUI/1.08"
            ),
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://fast.com/",
            "Origin": "https://fast.com",
        }
        request = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(request, timeout=15) as response:
            return json.loads(response.read().decode("utf-8"))

    def fastdotcom_measure_ping(self, url):
        """Measure ping by timing small requests to a fast.com target."""
        samples = []
        for _ in range(4):
            if self.stop_requested:
                return 0
            try:
                start = time.perf_counter()
                request = urllib.request.Request(url, method="GET")
                request.add_header("Range", "bytes=0-0")
                with urllib.request.urlopen(request, timeout=10) as response:
                    response.read()
                samples.append((time.perf_counter() - start) * 1000)
            except Exception:
                continue
        return min(samples) if samples else 0

    def fastdotcom_download_stream(self, url, byte_count, results, lock, stop_event):
        """Download from a single fast.com target, recording byte/time stats."""
        try:
            request = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 SpeedTestGUI/1.08",
                "Accept": "*/*",
                "Referer": "https://fast.com/",
            })
            with urllib.request.urlopen(request, timeout=self.FAST_STREAM_TIMEOUT) as response:
                start = time.perf_counter()
                downloaded = 0
                while downloaded < byte_count and not stop_event.is_set():
                    chunk = response.read(65536)
                    if not chunk:
                        break
                    downloaded += len(chunk)
                end = time.perf_counter()
            with lock:
                results.append({"bytes": downloaded, "start": start, "end": end})
        except Exception as e:
            with lock:
                results.append({"bytes": 0, "start": 0, "end": 0, "error": str(e)})

    def measure_fastdotcom(self):
        """Run a Netflix Fast.com measurement using parallel CDN streams."""
        if self.stop_requested:
            raise Exception("Test stopped by user")

        self.log_output("Fetching Fast.com targets from api.fast.com...")
        meta = self.fastdotcom_get_targets()
        targets = meta.get("targets", [])[:self.FAST_PARALLEL_STREAMS]
        if not targets:
            raise RuntimeError("Fast.com returned no targets")

        client = meta.get("client", {})
        client_isp = client.get("isp", "N/A")
        client_location = client.get("location", {})
        first_target_location = targets[0].get("location", {})
        server = "{city}, {country}".format(
            city=first_target_location.get("city", "?"),
            country=first_target_location.get("country", "?"),
        )
        self.log_output(f"Fast.com client ISP: {client_isp}")
        self.log_output(f"Fast.com client location: {client_location.get('city', '?')}, {client_location.get('country', '?')}")
        self.log_output(f"Nearest CDN target: {server}")

        self.log_output(f"Measuring ping against first target...")
        ping = self.fastdotcom_measure_ping(targets[0]["url"])
        self.log_output(f"Ping: {ping:.2f} ms")

        if self.stop_requested:
            raise Exception("Test stopped by user")

        self.log_output(
            f"Downloading {self.FAST_PARALLEL_STREAMS} parallel streams "
            f"of {self.FAST_SAMPLE_BYTES / 1_000_000:.0f} MB each..."
        )
        stream_results = []
        lock = threading.Lock()
        stop_event = threading.Event()
        threads = []
        for target in targets:
            thread = threading.Thread(
                target=self.fastdotcom_download_stream,
                args=(target["url"], self.FAST_SAMPLE_BYTES, stream_results, lock, stop_event),
            )
            thread.daemon = True
            thread.start()
            threads.append(thread)

        deadline = time.time() + self.FAST_STREAM_TIMEOUT + 10
        while any(t.is_alive() for t in threads):
            if self.stop_requested or time.time() > deadline:
                stop_event.set()
                break
            time.sleep(0.2)
        for thread in threads:
            thread.join(timeout=2)

        if self.stop_requested:
            raise Exception("Test stopped by user")

        valid = [r for r in stream_results if r["bytes"] > 0]
        if not valid:
            errors = [r.get("error", "no bytes") for r in stream_results]
            raise RuntimeError(f"All Fast.com streams failed: {errors}")

        earliest_start = min(r["start"] for r in valid)
        latest_end = max(r["end"] for r in valid)
        wall_clock = max(latest_end - earliest_start, 0.001)
        total_bytes = sum(r["bytes"] for r in valid)
        download_mbps = (total_bytes * 8) / wall_clock / 1_000_000

        self.log_output(
            f"Aggregated {total_bytes / 1_000_000:.1f} MB across {len(valid)} streams "
            f"in {wall_clock:.2f} s"
        )
        self.log_output(f"Download: {download_mbps:.2f} Mbps")
        self.log_output("Note: Fast.com does not support upload via this API.")

        return {
            "provider": "Fast.com",
            "download_mbps": download_mbps,
            "upload_mbps": None,
            "ping_ms": ping,
            "jitter_ms": None,
            "server": server,
            "isp": client_isp,
        }

    def measure_ookla_cli(self):
        """Run the Ookla speedtest CLI if installed; return None if absent."""
        speedtest_path = shutil.which("speedtest")
        if not speedtest_path:
            return None

        self.log_output(f"Found Ookla CLI at: {speedtest_path}")
        self.log_output("Running speedtest... (this can take up to 2 minutes)")

        try:
            self.ookla_process = subprocess.Popen(
                [
                    speedtest_path,
                    "--format=json",
                    "--accept-license",
                    "--accept-gdpr",
                    "--progress=no",
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )

            start_time = time.time()
            while self.ookla_process.poll() is None:
                if self.stop_requested:
                    self.ookla_process.terminate()
                    try:
                        self.ookla_process.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        self.ookla_process.kill()
                    return {"provider": "Ookla", "error": "Stopped by user"}
                if (time.time() - start_time) > self.OOKLA_TIMEOUT_SECONDS:
                    self.ookla_process.terminate()
                    try:
                        self.ookla_process.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        self.ookla_process.kill()
                    return {"provider": "Ookla", "error": f"Timed out after {self.OOKLA_TIMEOUT_SECONDS} s"}
                time.sleep(0.5)

            stdout, stderr = self.ookla_process.communicate()
            return_code = self.ookla_process.returncode

            if return_code != 0:
                error_message = (stderr or stdout or "").strip().splitlines()
                snippet = error_message[-1] if error_message else f"Exit code {return_code}"
                return {"provider": "Ookla", "error": snippet}

            data = json.loads(stdout)
            download_bw = data.get("download", {}).get("bandwidth", 0)
            upload_bw = data.get("upload", {}).get("bandwidth", 0)
            ping_data = data.get("ping", {})
            server_data = data.get("server", {})

            server_name = server_data.get("name", "Unknown")
            server_location = server_data.get("location", "")
            server_country = server_data.get("country", "")
            server_label = f"{server_name} ({server_location}, {server_country})".strip(" ,")

            download_mbps = download_bw * 8 / 1_000_000
            upload_mbps = upload_bw * 8 / 1_000_000
            ping_ms = ping_data.get("latency", 0)
            jitter_ms = ping_data.get("jitter", 0)

            self.log_output(f"Ookla server: {server_label}")
            self.log_output(f"Download: {download_mbps:.2f} Mbps")
            self.log_output(f"Upload: {upload_mbps:.2f} Mbps")
            self.log_output(f"Ping: {ping_ms:.2f} ms, Jitter: {jitter_ms:.2f} ms")

            return {
                "provider": "Ookla",
                "download_mbps": download_mbps,
                "upload_mbps": upload_mbps,
                "ping_ms": ping_ms,
                "jitter_ms": jitter_ms,
                "server": server_label,
                "isp": data.get("isp", "N/A"),
                "result_url": data.get("result", {}).get("url", ""),
            }
        except json.JSONDecodeError as e:
            return {"provider": "Ookla", "error": f"Could not parse JSON output: {e}"}
        except Exception as e:
            return {"provider": "Ookla", "error": str(e)}
        finally:
            self.ookla_process = None

    @staticmethod
    def _format_metric(value, unit, missing="N/A"):
        """Format a numeric metric for the comparison table."""
        if value is None:
            return missing
        try:
            return f"{value:.2f} {unit}"
        except (TypeError, ValueError):
            return missing

    def show_multi_provider_comparison(self, results):
        """Render the multi-provider comparison and spread analysis."""
        self.log_output("\n" + "=" * 80)
        self.log_output("MULTI-PROVIDER COMPARISON")
        self.log_output("=" * 80)

        header = f"{'Provider':<12}{'Download':>14}{'Upload':>14}{'Ping':>12}{'Jitter':>12}  Server"
        self.log_output(header)
        self.log_output("-" * 80)

        valid_downloads = []
        for result in results:
            if "error" in result:
                self.log_output(f"{result['provider']:<12}  ERROR: {result['error']}")
                continue

            download_label = self._format_metric(result.get("download_mbps"), "Mbps")
            upload_label = self._format_metric(result.get("upload_mbps"), "Mbps")
            ping_label = self._format_metric(result.get("ping_ms"), "ms")
            jitter_label = self._format_metric(result.get("jitter_ms"), "ms")
            server = result.get("server", "N/A")

            self.log_output(
                f"{result['provider']:<12}"
                f"{download_label:>14}"
                f"{upload_label:>14}"
                f"{ping_label:>12}"
                f"{jitter_label:>12}  {server}"
            )

            if result.get("download_mbps"):
                valid_downloads.append((result["provider"], result["download_mbps"]))

        self.log_output("")
        if len(valid_downloads) < 2:
            self.log_output("Need at least 2 successful providers for spread analysis.")
            return

        downloads = [value for _, value in valid_downloads]
        spread = max(downloads) - min(downloads)
        slowest = min(valid_downloads, key=lambda item: item[1])
        fastest = max(valid_downloads, key=lambda item: item[1])
        ratio = fastest[1] / slowest[1] if slowest[1] > 0 else 0

        self.log_output(
            f"Download spread: {spread:.2f} Mbps "
            f"({slowest[0]} {slowest[1]:.2f} -> {fastest[0]} {fastest[1]:.2f})"
        )
        self.log_output(f"Ratio fastest / slowest: {ratio:.2f}x")
        self.log_output("")

        if ratio >= 2.0:
            self.log_output("[!] Large spread between providers.")
            self.log_output("    Your ISP line is probably NOT the bottleneck.")
            self.log_output(
                f"    Traffic to {slowest[0]} is being throttled, mis-routed, "
                "or hitting a congested peering link."
            )
            self.log_output(
                f"    Real-world ceiling is closer to {fastest[1]:.2f} Mbps "
                f"(seen on {fastest[0]})."
            )
        elif ratio >= 1.3:
            self.log_output("[*] Moderate spread between providers.")
            self.log_output(
                "    Some path-specific variation exists, but no single provider "
                "is dramatically worse."
            )
        else:
            self.log_output("[OK] All providers agree within 30%.")
            self.log_output(
                f"    Real-world internet speed is approximately {min(downloads):.2f}"
                f" - {max(downloads):.2f} Mbps."
            )

    def save_multi_provider_log(self, results):
        """Append a multi-provider run to multi_provider_log.csv."""
        filename = os.path.join(os.path.dirname(__file__), self.MULTI_PROVIDER_LOG)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        header = [
            'timestamp', 'provider', 'download_mbps', 'upload_mbps',
            'ping_ms', 'jitter_ms', 'server', 'isp', 'error',
        ]
        file_exists = os.path.exists(filename)
        try:
            with open(filename, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(header)
                for result in results:
                    writer.writerow([
                        timestamp,
                        result.get('provider', ''),
                        result.get('download_mbps', ''),
                        result.get('upload_mbps', ''),
                        result.get('ping_ms', ''),
                        result.get('jitter_ms', ''),
                        result.get('server', ''),
                        result.get('isp', ''),
                        result.get('error', ''),
                    ])
            self.log_output(f"\nMulti-provider results appended to {filename}")
        except Exception as e:
            self.log_output(f"\nFailed to write multi-provider log: {e}")

    # =========================================================================
    # END MULTI-PROVIDER TESTING
    # =========================================================================

    def save_log(self):
        """Save the latest results, newest row at top."""
        if not self.latest_results:
            self.log_output("No test results available to log.")
            return

        comment = self.comment_entry.get().strip()
        if not comment:
            messagebox.showwarning(
                "Comment Required",
                "Please enter a short comment before saving to the log."
            )
            self.comment_entry.focus_set()
            return

        self.latest_results['comments'] = comment

        log_dir = os.path.dirname(__file__)
        filename = os.path.join(log_dir, 'internet_speed_log.csv')
        excel_filename = os.path.join(log_dir, 'internet_speed_log.xlsx')
        header = ['timestamp', 'server', 'sponsor', 'location', 'country', 'cc', 'lat', 'lon', 'isp', 'ip', 'download_mbps', 'upload_mbps', 'ping_ms', 'jitter_ms', 'comments']
        new_row = [
            self.latest_results['timestamp'],
            self.latest_results['server'],
            self.latest_results['sponsor'],
            self.latest_results['location'],
            self.latest_results['country'],
            self.latest_results['cc'],
            self.latest_results['lat'],
            self.latest_results['lon'],
            self.latest_results['isp'],
            self.latest_results['ip'],
            self.latest_results['download_mbps'],
            self.latest_results['upload_mbps'],
            self.latest_results['ping_ms'],
            self.latest_results['jitter_ms'],
            self.latest_results['comments']
        ]

        existing = []
        if os.path.exists(filename):
            try:
                with open(filename, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    existing = list(reader)
            except Exception:
                existing = []

        try:
            rows = [new_row]
            if existing and existing[0] == header:
                rows.extend(existing[1:])
            else:
                rows.extend(existing)

            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(rows)

            self.log_output(f"Logged results to {filename}")
        except Exception as e:
            self.log_output(f"Failed to write log: {e}")
            return

        try:
            self.write_excel_log(excel_filename, header, rows)
            self.log_output(f"Formatted Excel log saved to {excel_filename}")
        except Exception as e:
            self.log_output(f"Failed to write formatted Excel log: {e}")


def main():
    root = tk.Tk()
    app = SpeedTestApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()