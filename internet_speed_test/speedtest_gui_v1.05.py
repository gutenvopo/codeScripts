"""
Speed Test GUI Application v1.05
================================================================================
CHANGELOG - Version 1.05
================================================================================
V1.05 CHANGES:
1. Duplicated v1.04 into a new Cloudflare-backed version.
2. Replaced the `speedtest` library with direct calls to Cloudflare Speed Test endpoints.
3. Uses Cloudflare `/meta` for connection/IP/ISP details, `__down` for download,
   `__up` for upload, and lightweight requests for latency/jitter.
4. Keeps the existing Tkinter interface and CSV logging format.

V1.04 CHANGES:
1. Enhanced error handling for all HTTP error codes (not just 403)
2. Added extract_error_code() method to parse HTTP status codes from exceptions
3. Improved error reporting to include detailed server information:
   - Server sponsor/name
   - Location and country codes
   - Coordinates (latitude/longitude)
   - Full error details
4. Server details now displayed when ANY error occurs (not just 403)
5. Better user feedback during server retry operations
6. TODO: Verify correct CSV output/update flow; investigate why `internet_speed_log.csv` is not updating.

PREVIOUS (V1.03) CHANGES:
1. FIRST RELEASE BUILD - Production-ready version
2. Packaged as standalone executable with digital signature
3. Includes comprehensive speed testing with 403 error retry logic
4. Full CSV logging and comment support for test results

PREVIOUS (V1.02) FEATURES:
1. Added automatic retry mechanism for 403 Forbidden errors from servers
2. Displays which server caused the 403 error before retrying
3. Retries up to 3 times to find an alternate working server
4. Logs all 403 errors with server sponsor and location details

PREVIOUS (V1.01) FEATURES:
1. Added purple "Log" button to save test results to CSV file
2. Implemented CSV logging to 'internet_speed_log.csv' with full data export
3. CSV includes columns: timestamp, server, sponsor, location, country, cc, lat, 
   lon, isp, ip, download_mbps, upload_mbps, ping_ms, jitter_ms, comments
4. Newest test results are prepended to CSV (top row always latest)
5. Added comment input field with label "Comments, Keep it Short"
6. Comment input box spans full horizontal width of window
7. Log button disabled until test completes; enabled after results available
8. Window size increased to 900x620 to accommodate UI elements
9. Comments from input field are automatically saved with each log entry

WHAT THIS CODE DOES
1. Provides a desktop Tkinter GUI to run internet speed tests using Cloudflare Speed Test endpoints.
2. Displays real-time progress and detailed output in a scrollable text panel.
3. Collects and shows key metrics: download speed, upload speed, ping, and jitter.
4. Captures Cloudflare edge and network details (colo, location, country, coordinates, ISP, and IP).
5. Handles HTTP/network errors with clearer diagnostics.
6. Allows users to add short comments and log test results to `internet_speed_log.csv`.
7. Writes CSV output with the newest result at the top so recent tests are easy to find.

================================================================================
"""

import tkinter as tk
from tkinter import messagebox, scrolledtext
import threading
from datetime import datetime
import csv
import json
import os
import socket
import ssl
import sys
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Table = None
    TableStyleInfo = None
    get_column_letter = None

# Fix for PyInstaller GUI mode - redirect stdout/stderr
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')


class SpeedTestApp:
    CLOUDFLARE_BASE_URL = "https://speed.cloudflare.com"
    UPLOAD_SAMPLE_SIZES = [500_000, 1_000_000, 2_500_000, 5_000_000]
    UPLOAD_RETRIES = 2
    DARK_THEME = {
        "window": "#12161f",
        "panel": "#1a2030",
        "field": "#0f131b",
        "text": "#e8edf7",
        "muted": "#a9b4c7",
        "border": "#2a3346",
        "select_bg": "#2f6fed",
        "select_fg": "#ffffff",
        "disabled": "#7f8797",
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Speed Test Application v1.05 - Cloudflare")
        self.root.geometry("900x620")

        # Output text area
        self.output_frame = tk.Frame(root)
        self.output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.output_text = scrolledtext.ScrolledText(
            self.output_frame,
            wrap=tk.WORD,
            font=("Courier", 10),
            height=28,
            width=100
        )
        self.output_text.pack(fill=tk.BOTH, expand=True)
        self.output_text.insert(tk.END, "Click 'Start Test' to begin speed testing...\n")

        # Comment entry (above buttons)
        self.comment_frame = tk.Frame(root)
        self.comment_frame.pack(fill=tk.X, padx=10, pady=(0, 6))
        self.comment_label = tk.Label(self.comment_frame, text="Comments, Keep it Short", font=("Arial", 10))
        self.comment_label.pack(anchor='w')
        self.comment_entry = tk.Entry(self.comment_frame)
        self.comment_entry.pack(fill=tk.X, expand=True, padx=0, pady=(4, 0))

        # Button frame
        self.button_frame = tk.Frame(root)
        self.button_frame.pack(fill=tk.X, padx=10, pady=6)

        # Start Test button
        self.start_button = tk.Button(
            self.button_frame,
            text="Start Test",
            command=self.start_test,
            width=15,
            height=2,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.start_button.pack(side=tk.LEFT, padx=5)

        # Re-test button
        self.retest_button = tk.Button(
            self.button_frame,
            text="Re-test",
            command=self.start_test,
            width=15,
            height=2,
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.retest_button.pack(side=tk.LEFT, padx=5)

        # Stop button (different color)
        self.stop_button = tk.Button(
            self.button_frame,
            text="Stop",
            command=self.stop_test,
            width=15,
            height=2,
            bg="#FF9800",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)

        # Log button (different color)
        self.log_button = tk.Button(
            self.button_frame,
            text="Save To Log",
            command=self.save_log,
            width=15,
            height=2,
            bg="#9C27B0",
            fg="white",
            font=("Arial", 10, "bold"),
            disabledforeground="grey"
        )
        self.log_button.pack(side=tk.LEFT, padx=5)
        self.log_button.config(state=tk.DISABLED)  # disabled until results exist

        # Open Log button
        self.open_log_button = tk.Button(
            self.button_frame,
            text="Open Log",
            command=self.open_log,
            width=15,
            height=2,
            bg="#607D8B",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.open_log_button.pack(side=tk.RIGHT, padx=5)

        # Exit button
        self.exit_button = tk.Button(
            self.button_frame,
            text="Exit",
            command=self.root.quit,
            width=15,
            height=2,
            bg="#f44336",
            fg="white",
            font=("Arial", 10, "bold")
        )
        self.exit_button.pack(side=tk.LEFT, padx=5)

        self.test_running = False
        self.stop_requested = False
        self.latest_results = None
        self.tls_warning_shown = False
        self.apply_dark_mode()

    def apply_dark_mode(self):
        """Apply a dark color theme to the Tkinter controls."""
        theme = self.DARK_THEME
        self.root.configure(bg=theme["window"])

        for frame in (self.output_frame, self.comment_frame, self.button_frame):
            frame.configure(bg=theme["window"])

        self.comment_label.configure(
            bg=theme["window"],
            fg=theme["muted"],
            activebackground=theme["window"],
            activeforeground=theme["text"],
        )
        self.comment_entry.configure(
            bg=theme["field"],
            fg=theme["text"],
            insertbackground=theme["text"],
            selectbackground=theme["select_bg"],
            selectforeground=theme["select_fg"],
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=theme["border"],
            highlightcolor=theme["select_bg"],
        )
        self.output_text.configure(
            bg=theme["field"],
            fg=theme["text"],
            insertbackground=theme["text"],
            selectbackground=theme["select_bg"],
            selectforeground=theme["select_fg"],
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=theme["border"],
            highlightcolor=theme["select_bg"],
        )

        for button in (
            self.start_button,
            self.retest_button,
            self.stop_button,
            self.log_button,
            self.open_log_button,
            self.exit_button,
        ):
            button.configure(
                fg="white",
                activeforeground="white",
                disabledforeground=theme["disabled"],
                relief=tk.FLAT,
                borderwidth=0,
                highlightthickness=0,
                cursor="hand2",
            )

        self.start_button.configure(activebackground="#3f9a44")
        self.retest_button.configure(activebackground="#1976d2")
        self.stop_button.configure(activebackground="#ef8f00")
        self.log_button.configure(activebackground="#8e24aa")
        self.open_log_button.configure(activebackground="#546e7a")
        self.exit_button.configure(activebackground="#d83a30")

    def extract_error_code(self, error_string):
        """Extract HTTP error code from error message"""
        import re
        # Look for patterns like '403', 'HTTP 403', 'Status: 403'
        match = re.search(r'(\d{3})', error_string)
        if match:
            return match.group(1)
        return 'UNKNOWN'

    def log_output(self, message):
        """Add message to output text area"""
        self.output_text.insert(tk.END, message + "\n")
        self.output_text.see(tk.END)
        self.root.update()

    def log_verbose_error(self, error):
        """Print detailed error diagnostics to the GUI output."""
        self.log_output("\n" + "=" * 60)
        self.log_output("VERBOSE ERROR DETAILS")
        self.log_output("=" * 60)
        self.log_output(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_output(f"Error Type: {type(error).__name__}")
        self.log_output(f"Error Message: {error}")

        cause = error.__cause__ or error.__context__
        if cause:
            self.log_output(f"Underlying Type: {type(cause).__name__}")
            self.log_output(f"Underlying Message: {cause}")

        self.log_output("")
        self.log_output("Traceback:")
        for line in traceback.format_exception(type(error), error, error.__traceback__):
            for output_line in line.rstrip().splitlines():
                self.log_output(output_line)
        self.log_output("=" * 60)

    def open_log(self):
        """Open the Excel log file in the default system app."""
        log_dir = os.path.dirname(__file__)
        filename = os.path.join(log_dir, 'internet_speed_log.xlsx')
        if not os.path.exists(filename):
            filename = os.path.join(log_dir, 'internet_speed_log.csv')

        if not os.path.exists(filename):
            self.log_output(f"Log file not found yet: {filename}")
            self.log_output("Run a test and click 'Save To Log' first.")
            return

        try:
            os.startfile(filename)
        except Exception as e:
            self.log_output(f"Failed to open log: {e}")

    def write_excel_log(self, filename, header, rows):
        """Write log rows as a formatted Excel table."""
        if Workbook is None:
            raise ImportError("openpyxl is required to write formatted Excel logs")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Speed Test Log"

        worksheet.append(header)
        for row in rows:
            worksheet.append(row)

        table_end_column = get_column_letter(len(header))
        table_end_row = max(len(rows) + 1, 2)
        table_ref = f"A1:{table_end_column}{table_end_row}"
        table = Table(displayName="SpeedTestLog", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium13",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

        for column_index, column_name in enumerate(header, start=1):
            values = [column_name]
            values.extend(str(row[column_index - 1]) for row in rows if len(row) >= column_index)
            width = min(max(len(value) for value in values) + 2, 45)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        workbook.save(filename)

    def stop_test(self):
        """Stop the running test"""
        if not self.test_running:
            self.output_text.delete(1.0, tk.END)
            self.log_output("No tests are running")
            self.log_output("")
            self.log_output("Click 'Start Test' to begin speed testing...\n")
            return
        
        self.stop_requested = True
        self.log_output("\n[INFO] User stopped the tests")
        self.log_output("")
        self.log_output("Click 'Start Test' to begin speed testing...\n")

    def draw_diagonal_lines(self):
        """Draw diagonal lines on canvas for disabled state"""
        self.log_button_canvas.delete("all")
        width = 100
        height = 48
        spacing = width // 6
        for i in range(8):
            x_start = i * spacing - height
            self.log_button_canvas.create_line(
                x_start, 0,
                x_start + height, height,
                fill="grey",
                width=2
            )

    def start_test(self):
        """Start the speed test in a separate thread"""
        if self.test_running:
            self.log_output("Test already in progress...")
            return

        self.test_running = True
        self.start_button.config(state=tk.DISABLED)
        self.retest_button.config(state=tk.DISABLED)
        self.log_button.config(state=tk.DISABLED)

        self.output_text.delete(1.0, tk.END)
        self.log_output("=" * 80)
        self.log_output(f"Speed Test Started - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log_output("=" * 80)
        self.log_output("")

        # Run speed test in separate thread
        test_thread = threading.Thread(target=self.run_speedtest)
        test_thread.daemon = True
        test_thread.start()

    def build_cloudflare_request(self, path, params=None, data=None, method=None):
        """Build a Cloudflare Speed Test request with browser-like headers."""
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36 SpeedTestGUI/1.05"
            ),
            "Accept": "application/json,text/plain,*/*",
            "Referer": f"{self.CLOUDFLARE_BASE_URL}/",
            "Origin": self.CLOUDFLARE_BASE_URL,
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }
        if data is not None:
            headers["Content-Type"] = "application/octet-stream"

        query = ""
        if params:
            query = "?" + urllib.parse.urlencode(params)

        url = f"{self.CLOUDFLARE_BASE_URL}{path}{query}"
        return urllib.request.Request(url, data=data, headers=headers, method=method)

    def open_cloudflare_request(self, request, timeout, context=None):
        """Open a Cloudflare request and normalize HTTP errors."""
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
                return response.read(), response
        except urllib.error.HTTPError as e:
            raise Exception(f"HTTP {e.code} from Cloudflare: {e.reason}") from e
        except (TimeoutError, socket.timeout) as e:
            raise Exception(f"Network timeout calling Cloudflare: {e}") from e

    def cloudflare_request(self, path, params=None, data=None, method=None, timeout=30):
        """Make one request to Cloudflare Speed Test."""
        if self.stop_requested:
            raise Exception("Test stopped by user")

        request_method = method or ("POST" if data is not None else "GET")
        payload_size = len(data) if data is not None else 0
        request_details = (
            f"{request_method} {path}, timeout={timeout}s, "
            f"payload={payload_size / 1_000_000:.2f} MB"
        )
        request = self.build_cloudflare_request(path, params, data, method)

        try:
            return self.open_cloudflare_request(request, timeout)
        except urllib.error.URLError as e:
            if isinstance(e.reason, ssl.SSLCertVerificationError):
                if not self.tls_warning_shown:
                    self.log_output("[WARN] TLS certificate verification failed; retrying without certificate verification.")
                    self.tls_warning_shown = True
                context = ssl._create_unverified_context()
                request = self.build_cloudflare_request(path, params, data, method)
                return self.open_cloudflare_request(request, timeout, context)
            raise Exception(f"Network error calling Cloudflare ({request_details}): {e.reason}") from e
        except (TimeoutError, socket.timeout) as e:
            raise Exception(f"Network timeout calling Cloudflare ({request_details}): {e}") from e

    def get_cloudflare_meta(self):
        """Fetch Cloudflare metadata for client and edge details."""
        try:
            body, _ = self.cloudflare_request("/meta", timeout=15)
            return json.loads(body.decode("utf-8"))
        except Exception as e:
            self.log_output(f"[WARN] Cloudflare /meta failed ({e}); using /cdn-cgi/trace fallback.")
            body, _ = self.cloudflare_request("/cdn-cgi/trace", timeout=15)
            trace = {}
            for line in body.decode("utf-8", errors="replace").splitlines():
                if "=" in line:
                    key, value = line.split("=", 1)
                    trace[key] = value
            return {
                "hostname": trace.get("h", "speed.cloudflare.com"),
                "clientIp": trace.get("ip", "N/A"),
                "httpProtocol": trace.get("http", "N/A"),
                "asOrganization": "N/A",
                "country": trace.get("loc", "N/A"),
                "city": "N/A",
                "region": "N/A",
                "latitude": "N/A",
                "longitude": "N/A",
                "colo": trace.get("colo", "Cloudflare"),
            }

    def measure_latency(self, samples=6):
        """Measure latency and jitter with lightweight Cloudflare requests."""
        latencies = []
        self.log_output("Calculating ping and jitter...")

        for i in range(samples):
            if self.stop_requested:
                raise Exception("Test stopped by user")
            start = time.perf_counter()
            self.cloudflare_request(
                "/__down",
                params={"bytes": 0, "cachebust": f"{time.time_ns()}-{i}"},
                timeout=10
            )
            elapsed_ms = (time.perf_counter() - start) * 1000
            latencies.append(elapsed_ms)

        ping = min(latencies)
        jitter = max(latencies) - min(latencies) if len(latencies) > 1 else 0.0
        return ping, jitter

    def measure_download(self):
        """Measure download speed through Cloudflare __down."""
        total_bytes = 0
        total_seconds = 0.0
        sizes = [1_000_000, 5_000_000, 10_000_000, 25_000_000]

        for size in sizes:
            if self.stop_requested:
                raise Exception("Test stopped by user")
            self.log_output(f"Downloading {size / 1_000_000:.0f} MB sample...")
            start = time.perf_counter()
            body, _ = self.cloudflare_request(
                "/__down",
                params={"bytes": size, "cachebust": time.time_ns()},
                timeout=45
            )
            elapsed = time.perf_counter() - start
            total_bytes += len(body)
            total_seconds += elapsed

        return (total_bytes * 8) / total_seconds / 1_000_000

    def measure_upload(self):
        """Measure upload speed through Cloudflare __up."""
        total_bytes = 0
        total_seconds = 0.0

        for size in self.UPLOAD_SAMPLE_SIZES:
            if self.stop_requested:
                raise Exception("Test stopped by user")
            sample_mb = size / 1_000_000
            payload = b"0" * size

            for attempt in range(1, self.UPLOAD_RETRIES + 2):
                try:
                    self.log_output(f"Uploading {sample_mb:.1f} MB sample...")
                    start = time.perf_counter()
                    self.cloudflare_request("/__up", data=payload, method="POST", timeout=60)
                    elapsed = time.perf_counter() - start
                    total_bytes += size
                    total_seconds += elapsed
                    break
                except Exception as e:
                    if self.stop_requested or "Test stopped by user" in str(e):
                        raise
                    if attempt > self.UPLOAD_RETRIES:
                        if total_bytes:
                            self.log_output(f"[WARN] Skipping {sample_mb:.1f} MB upload sample after timeout/error: {e}")
                            break
                        raise Exception(
                            "Upload test could not reach Cloudflare. "
                            f"Last error: {e}"
                        ) from e
                    self.log_output(
                        f"[WARN] Upload sample failed ({e}); retrying "
                        f"{attempt}/{self.UPLOAD_RETRIES}..."
                    )

        if total_seconds <= 0:
            raise Exception("Upload test did not complete any samples.")

        return (total_bytes * 8) / total_seconds / 1_000_000

    def run_speedtest(self):
        """Run the Cloudflare speed test."""
        try:
            self.log_output("Initializing Cloudflare Speed Test...")
            self.log_output("Getting Cloudflare metadata...")
            meta = self.get_cloudflare_meta()

            location_parts = [
                meta.get("city"),
                meta.get("region"),
                meta.get("country"),
            ]
            location = ", ".join(part for part in location_parts if part) or "N/A"
            lat = meta.get("latitude", "N/A")
            lon = meta.get("longitude", "N/A")
            if meta.get("loc") and "," in meta["loc"]:
                lat, lon = [part.strip() for part in meta["loc"].split(",", 1)]

            colo = meta.get("colo", "Cloudflare")
            if isinstance(colo, dict):
                server_name = colo.get("iata", "Cloudflare")
            else:
                server_name = colo
            sponsor = "Cloudflare"
            isp = meta.get("asOrganization", "N/A")
            ip_address = meta.get("clientIp", "N/A")
            country = meta.get("country", "N/A")

            self.log_output("\n" + "=" * 60)
            self.log_output("CLOUDFLARE EDGE INFORMATION")
            self.log_output("=" * 60)
            self.log_output(f"Server: {server_name}")
            self.log_output(f"Sponsor: {sponsor}")
            self.log_output(f"Location: {location}")
            self.log_output(f"Country Code: {country}")
            self.log_output(f"Latitude: {lat}")
            self.log_output(f"Longitude: {lon}")

            self.log_output("\n" + "=" * 60)
            self.log_output("ISP INFORMATION")
            self.log_output("=" * 60)
            self.log_output(f"ISP: {isp}")
            self.log_output(f"IP Address: {ip_address}")
            self.log_output(f"Country: {country}")

            self.log_output("\n" + "=" * 60)
            self.log_output("PING & LATENCY")
            self.log_output("=" * 60)
            ping, jitter = self.measure_latency()
            self.log_output(f"Ping (Latency): {ping:.2f} ms")
            self.log_output(f"Jitter: {jitter:.2f} ms")

            self.log_output("\n" + "=" * 60)
            self.log_output("DOWNLOAD SPEED TEST")
            self.log_output("=" * 60)
            download_speed = self.measure_download()
            self.log_output(f"Download Speed: {download_speed:.2f} Mbps")

            self.log_output("\n" + "=" * 60)
            self.log_output("UPLOAD SPEED TEST")
            self.log_output("=" * 60)
            upload_speed = self.measure_upload()
            self.log_output(f"Upload Speed: {upload_speed:.2f} Mbps")

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.latest_results = {
                'timestamp': timestamp,
                'server': server_name,
                'sponsor': sponsor,
                'location': location,
                'country': country,
                'cc': country,
                'lat': lat,
                'lon': lon,
                'isp': isp,
                'ip': ip_address,
                'download_mbps': round(download_speed, 2),
                'upload_mbps': round(upload_speed, 2),
                'ping_ms': round(ping, 2),
                'jitter_ms': round(jitter, 2),
                'comments': ''
            }

            self.log_output("\n" + "=" * 60)
            self.log_output("TEST SUMMARY")
            self.log_output("=" * 60)
            self.log_output(f"Download: {download_speed:.2f} Mbps")
            self.log_output(f"Upload: {upload_speed:.2f} Mbps")
            self.log_output(f"Ping: {ping:.2f} ms")
            self.log_output(f"Jitter: {jitter:.2f} ms")
            self.log_output("=" * 60)
            self.log_output(f"Test Completed - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self.log_output("=" * 60)

            self.log_button.config(state=tk.NORMAL)

        except Exception as e:
            self.log_output(f"\nERROR: {str(e)}")
            if "Test stopped by user" not in str(e):
                self.log_verbose_error(e)
                self.log_output("Please check your internet connection and try again.")

        finally:
            self.test_running = False
            self.stop_requested = False
            self.start_button.config(state=tk.NORMAL)
            self.retest_button.config(state=tk.NORMAL)

    def save_log(self):
        """Save the latest results, newest row at top."""
        if not self.latest_results:
            self.log_output("No test results available to log.")
            return

        # Attach comment from entry
        comment = self.comment_entry.get().strip()
        if not comment:
            messagebox.showwarning(
                "Comment Required",
                "Please enter a short comment before saving to the log."
            )
            self.comment_entry.focus_set()
            return

        self.latest_results['comments'] = comment

        log_dir = os.path.dirname(__file__)
        filename = os.path.join(log_dir, 'internet_speed_log.csv')
        excel_filename = os.path.join(log_dir, 'internet_speed_log.xlsx')
        header = ['timestamp', 'server', 'sponsor', 'location', 'country', 'cc', 'lat', 'lon', 'isp', 'ip', 'download_mbps', 'upload_mbps', 'ping_ms', 'jitter_ms', 'comments']
        new_row = [
            self.latest_results['timestamp'],
            self.latest_results['server'],
            self.latest_results['sponsor'],
            self.latest_results['location'],
            self.latest_results['country'],
            self.latest_results['cc'],
            self.latest_results['lat'],
            self.latest_results['lon'],
            self.latest_results['isp'],
            self.latest_results['ip'],
            self.latest_results['download_mbps'],
            self.latest_results['upload_mbps'],
            self.latest_results['ping_ms'],
            self.latest_results['jitter_ms'],
            self.latest_results['comments']
        ]

        # Read existing rows
        existing = []
        if os.path.exists(filename):
            try:
                with open(filename, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    existing = list(reader)
            except Exception:
                existing = []

        # Write header and new row, then existing rows (so newest at top)
        try:
            rows = [new_row]
            if existing and existing[0] == header:
                rows.extend(existing[1:])
            else:
                rows.extend(existing)

            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(rows)

            self.log_output(f"Logged results to {filename}")
        except Exception as e:
            self.log_output(f"Failed to write log: {e}")
            return

        try:
            self.write_excel_log(excel_filename, header, rows)
            self.log_output(f"Formatted Excel log saved to {excel_filename}")
        except Exception as e:
            self.log_output(f"Failed to write formatted Excel log: {e}")


def main():
    root = tk.Tk()
    app = SpeedTestApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
